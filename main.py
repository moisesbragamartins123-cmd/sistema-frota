import streamlit as st
from supabase import create_client
import pandas as pd
from datetime import datetime, date
import plotly.express as px
import plotly.graph_objects as go
import os
import time
import io
from fpdf import FPDF

# ═══════════════════════════════════════════════════════════════════
# CONFIGURAÇÃO DA PÁGINA
# ═══════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="PavControl — COPA Engenharia",
    page_icon="🛣️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS GERAL DA APLICAÇÃO
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
.main { background-color: #F4F7FB; }
[data-testid="stSidebar"] { background: #0F1923; }
[data-testid="stSidebar"] * { color: #C9D4E0 !important; }
[data-testid="stSidebar"] h3 { color: #1D9E75 !important; }

/* ═════════ ESTILO DOS CAMPOS DE PREENCHIMENTO E TEXTOS ═════════ */
.stTextInput>label, .stSelectbox>label, .stNumberInput>label,
.stDateInput>label, .stTimeInput>label, .stTextArea>label {
    font-size: 13px !important; 
    color: #1E293B !important; 
    font-weight: 700 !important; 
    text-transform: uppercase;
    letter-spacing: 0.05em;
    margin-bottom: 6px !important;
}
div[data-baseweb="input"], div[data-baseweb="select"] {
    border-radius: 8px !important;
    border: 1px solid #CBD5E1 !important;
    background-color: #F8FAFC !important; 
    padding: 2px !important;
}
div[data-baseweb="input"]:focus-within, div[data-baseweb="select"]:focus-within {
    border-color: #0A58CA !important;
    box-shadow: 0 0 0 3px rgba(10, 88, 202, 0.15) !important;
    background-color: #FFFFFF !important;
}

/* ═════════ ESTILO DOS BOTÕES ═════════ */
div.stButton > button:first-child {
    background-color: #0A58CA !important; 
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 800 !important;
    font-size: 15px !important; 
    padding: 0.75rem 1.5rem !important;
    box-shadow: 0 4px 6px rgba(0,0,0,0.15) !important;
    transition: all 0.2s ease-in-out !important;
    text-transform: uppercase !important;
}
div.stButton > button:first-child:hover {
    background-color: #084298 !important;
    transform: translateY(-2px) !important;
    box-shadow: 0 6px 12px rgba(0,0,0,0.2) !important;
}

/* ═════════ ESTILO DOS CARDS E AVISOS ═════════ */
div[data-testid="stForm"] {
    border: none; border-radius: 16px;
    padding: 2rem; background: white;
    box-shadow: 0 10px 30px rgba(0,0,0,0.08);
}
.banner-ok  { background:#EAF3DE; color:#3B6D11; border:1px solid #C0DD97; border-radius:10px; padding:12px 16px; font-weight:600; font-size:13px; margin-bottom:1rem; box-shadow: 0 2px 4px rgba(0,0,0,0.02); }
.banner-low { background:#FAEEDA; color:#854F0B; border:1px solid #FAC775; border-radius:10px; padding:12px 16px; font-weight:600; font-size:13px; margin-bottom:1rem; box-shadow: 0 2px 4px rgba(0,0,0,0.02); }
.banner-err { background:#FCEBEB; color:#A32D2D; border:1px solid #F0B0AE; border-radius:10px; padding:12px 16px; font-weight:600; font-size:13px; margin-bottom:1rem; box-shadow: 0 2px 4px rgba(0,0,0,0.02); }
.banner-info{ background:#E6F1FB; color:#185FA5; border:1px solid #A8C9EE; border-radius:10px; padding:12px 16px; font-weight:600; font-size:13px; margin-bottom:1rem; box-shadow: 0 2px 4px rgba(0,0,0,0.02); }
.kpi-box { background:white; border:none; border-radius:12px; padding:1.5rem; text-align:center; height:100%; box-shadow: 0 4px 15px rgba(0,0,0,0.04); transition: transform 0.2s ease; }
.kpi-box:hover { transform: translateY(-2px); box-shadow: 0 8px 20px rgba(0,0,0,0.08); }
.kpi-box h3 { margin:0; font-size:24px; color:#0F1923; margin-top: 10px; }
.kpi-box p  { margin:0; font-size:11px; color:#64748B; text-transform:uppercase; font-weight:700; letter-spacing: 0.05em; }
.tag-tanque { background:#DBEAFE; color:#1E40AF; border-radius:6px; padding:2px 8px; font-size:11px; font-weight:700; }
.tag-posto  { background:#D1FAE5; color:#065F46; border-radius:6px; padding:2px 8px; font-size:11px; font-weight:700; }
.tag-obra   { background:#FEF3C7; color:#92400E; border-radius:6px; padding:2px 8px; font-size:11px; font-weight:700; }
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════
# SUPABASE E BANCO DE DADOS
# ═══════════════════════════════════════════════════════════════════
@st.cache_resource
def get_supabase():
    return create_client(st.secrets["SUPABASE_URL"], st.secrets["SUPABASE_KEY"])

supabase = get_supabase()

def get_data(table: str) -> pd.DataFrame:
    try:
        res = supabase.table(table).select("*").execute()
        return pd.DataFrame(res.data) if res.data else pd.DataFrame()
    except Exception as e:
        st.warning(f"⚠️ Erro ao buscar {table}: {e}")
        return pd.DataFrame()

def insert_data(table: str, data: dict) -> bool:
    try:
        supabase.table(table).insert(data).execute()
        return True
    except Exception as e:
        st.error(f"❌ Erro ao salvar: {e}")
        return False

def update_data(table: str, row_id, data: dict) -> bool:
    try:
        supabase.table(table).update(data).eq("id", row_id).execute()
        return True
    except Exception as e:
        st.error(f"❌ Erro ao atualizar: {e}")
        return False

def delete_data(table: str, row_id) -> bool:
    try:
        supabase.table(table).delete().eq("id", row_id).execute()
        return True
    except Exception as e:
        st.error(f"❌ Erro ao excluir: {e}")
        return False

def calcular_saldo(nome_tanque: str) -> float:
    """
    Calcula o saldo de um tanque considerando:
    - Entradas: tabela entradas_tanque
    - Saídas diretas: abastecimentos com origem=Tanque Interno (status=ATIVO)
    - Saídas para caminhão-tanque: transferencias_tanque (status=ATIVO)
    """
    df_ent = get_data("entradas_tanque")
    df_sai = get_data("abastecimentos")
    df_transf = get_data("transferencias_tanque")

    # Apenas registros ativos
    if not df_sai.empty and "status" in df_sai.columns:
        df_sai = df_sai[df_sai["status"] == "ATIVO"]
    if not df_transf.empty and "status" in df_transf.columns:
        df_transf = df_transf[df_transf["status"] == "ATIVO"]

    t_ent = 0.0
    if not df_ent.empty and "nome_tanque" in df_ent.columns:
        t_ent = pd.to_numeric(
            df_ent[df_ent["nome_tanque"] == nome_tanque]["quantidade"], errors="coerce"
        ).sum()

    # Saídas diretas (abastecimentos de veículos a partir do tanque)
    t_sai_direto = 0.0
    if not df_sai.empty and "nome_tanque" in df_sai.columns and "origem" in df_sai.columns:
        mask = (df_sai["origem"] == "Tanque Interno") & (df_sai["nome_tanque"] == nome_tanque)
        t_sai_direto = pd.to_numeric(df_sai.loc[mask, "quantidade"], errors="coerce").sum()

    # Saídas para caminhão-tanque (transferências)
    t_sai_transf = 0.0
    if not df_transf.empty and "tanque_origem" in df_transf.columns:
        mask_t = df_transf["tanque_origem"] == nome_tanque
        t_sai_transf = pd.to_numeric(df_transf.loc[mask_t, "quantidade"], errors="coerce").sum()

    return float(t_ent) - float(t_sai_direto) - float(t_sai_transf)

def dia_semana_pt(d) -> str:
    dias = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"]
    try:
        if isinstance(d, str):
            d = datetime.strptime(d[:10], "%Y-%m-%d")
        return dias[d.weekday()]
    except:
        return ""

def lista_obras(incluir_todas: bool = False) -> list:
    """Retorna lista de obras cadastradas para uso em selectboxes."""
    df_o = get_data("obras")
    if df_o.empty:
        return ["GERAL"] if incluir_todas else []
    nomes = df_o[df_o.get("status", pd.Series(["Ativa"] * len(df_o))) != "Encerrada"]["nome"].tolist()
    if incluir_todas:
        return ["TODAS"] + nomes
    return nomes

# ═══════════════════════════════════════════════════════════════════
# EXPORTAÇÃO — EXCEL PADRÃO COPA E PDF TIMBRADO
# ═══════════════════════════════════════════════════════════════════
def gerar_excel_copa(df: pd.DataFrame, dados_forn: dict, periodo: str, obra: str, nome_forn: str = "") -> bytes:
    df = df.fillna("").copy()
    dias_pt = {0: "SEG", 1: "TER", 2: "QUA", 3: "QUI", 4: "SEX", 5: "SÁB", 6: "DOM"}
    template = "template_posto.xlsx"
    if os.path.exists(template):
        from openpyxl import load_workbook
        wb = load_workbook(template)
        ws = wb.active
        ws["D1"] = obra.upper()
        ws["D3"] = periodo.upper()
        ws["J1"] = dados_forn.get("razao_social", dados_forn.get("nome", "")).upper()
        ws["J2"] = dados_forn.get("agencia", "")
        ws["J3"] = dados_forn.get("conta", "")
        ws["M1"] = dados_forn.get("pix", "")
        ws["M2"] = dados_forn.get("tipo_conta", "")
        ws["M3"] = dados_forn.get("banco", "")
        row0 = 8
        for i, (_, r) in enumerate(df.iterrows()):
            dia_str = ""
            try:
                dia_str = dias_pt[datetime.strptime(str(r.get("data", ""))[:10], "%Y-%m-%d").weekday()]
            except:
                pass
            qtd  = float(r.get("quantidade",    0) or 0)
            vunt = float(r.get("valor_unitario", 0) or 0)
            tot  = float(r.get("total",          0) or 0)
            ws.cell(row0 + i, 1,  str(r.get("data", ""))[:10])
            ws.cell(row0 + i, 2,  dia_str)
            ws.cell(row0 + i, 3,  str(r.get("numero_ficha", "")))
            ws.cell(row0 + i, 4,  str(r.get("placa", "")))
            ws.cell(row0 + i, 5,  str(r.get("prefixo", "")))
            ws.cell(row0 + i, 6,  str(r.get("motorista", "")))
            ws.cell(row0 + i, 7,  str(r.get("fornecedor", "")))
            ws.cell(row0 + i, 8,  str(r.get("tipo_combustivel", "")))
            ws.cell(row0 + i, 9,  qtd).number_format = "#,##0.00"
            ws.cell(row0 + i, 10, vunt).number_format = '"R$" #,##0.00'
            ws.cell(row0 + i, 11, tot).number_format  = '"R$" #,##0.00'
            ws.cell(row0 + i, 12, str(r.get("horimetro", "")))
            ws.cell(row0 + i, 13, str(r.get("obra", "")))
            ws.cell(row0 + i, 14, str(r.get("observacao", "")))
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    import xlsxwriter
    buf = io.BytesIO()
    wb  = xlsxwriter.Workbook(buf)
    ws  = wb.add_worksheet("Abastecimento")
    fh  = wb.add_format({"bold": True, "font_size": 9, "border": 1, "align": "center", "valign": "vcenter", "bg_color": "#BDD7EE", "text_wrap": True})
    fd  = wb.add_format({"font_size": 9, "border": 1, "align": "center"})
    fn  = wb.add_format({"font_size": 9, "border": 1, "num_format": "#,##0.00", "align": "right"})
    fm  = wb.add_format({"font_size": 9, "border": 1, "num_format": '"R$" #,##0.00', "align": "right"})
    ft  = wb.add_format({"bold": True, "font_size": 9, "border": 1, "num_format": "#,##0.00", "align": "right", "bg_color": "#D9D9D9"})
    ftm = wb.add_format({"bold": True, "font_size": 9, "border": 1, "num_format": '"R$" #,##0.00', "align": "right", "bg_color": "#D9D9D9"})
    ftx = wb.add_format({"bold": True, "font_size": 9, "border": 1, "align": "right", "bg_color": "#D9D9D9"})
    flb = wb.add_format({"bold": True, "font_size": 9})
    fvl = wb.add_format({"font_size": 9})
    fti = wb.add_format({"bold": True, "font_size": 11, "align": "center", "bg_color": "#D9D9D9", "border": 1})

    ws.write(0, 0, "OBRA:",  flb); ws.write(0, 2, obra.upper(), fvl)
    ws.write(0, 5, "PREÇO GASOLINA (L)", flb); ws.write(0, 6, dados_forn.get("preco_gasolina", "") or "", fvl)
    ws.write(0, 8, "RAZÃO SOCIAL", flb); ws.write(0, 10, dados_forn.get("razao_social", ""), fvl)
    ws.write(0, 12, "TIPO DE CONTA:", flb); ws.write(0, 13, dados_forn.get("tipo_conta", ""), fvl)
    ws.write(1, 0, "DESCRIÇÃO:", flb)
    ws.write(1, 5, "PREÇO DIESEL (L)", flb); ws.write(1, 6, dados_forn.get("preco_diesel", "") or "", fvl)
    ws.write(1, 8, "AGÊNCIA:", flb); ws.write(1, 10, dados_forn.get("agencia", ""), fvl)
    ws.write(1, 12, "BANCO:", flb); ws.write(1, 13, dados_forn.get("banco", ""), fvl)
    ws.write(2, 0, "PERÍODO:", flb); ws.write(2, 2, periodo.upper(), fvl)
    ws.write(2, 8, "CONTA:", flb); ws.write(2, 10, dados_forn.get("conta", ""), fvl)
    ws.write(2, 12, "PIX:", flb); ws.write(2, 13, dados_forn.get("pix", ""), fvl)

    titulo = f"CONTROLE DE ABASTECIMENTO  —  {nome_forn.upper() or obra.upper()}"
    ws.merge_range(3, 0, 3, 14, titulo, fti)

    heads  = ["DATA", "DIA DA\nSEMANA", "FICHA", "PLACA", "CÓDIGO /\nPREFIXO",
              "VEÍCULO / MAQUINA -\nMOTORISTA / OPERADOR", "FORNECEDOR",
              "TIPO DE\nCOMBUSTÍVEL", "QUANTIDADE\n(L)", "VALOR\nUNITÁRIO (R$)",
              "TOTAL\n(R$)", "KM / HOR", "OBRA", "OBSERVAÇÃO"]
    widths = [12, 8, 8, 10, 10, 30, 20, 12, 10, 12, 12, 10, 20, 30]
    ws.set_row(4, 36)
    for ci, (h, w) in enumerate(zip(heads, widths)):
        ws.write(4, ci, h, fh)
        ws.set_column(ci, ci, w)

    t_l = 0.0; t_r = 0.0
    for ri, (_, r) in enumerate(df.iterrows(), start=5):
        dia_str = ""
        try:
            dia_str = dias_pt[datetime.strptime(str(r.get("data", ""))[:10], "%Y-%m-%d").weekday()]
        except:
            pass
        qtd  = float(r.get("quantidade",    0) or 0)
        vunt = float(r.get("valor_unitario", 0) or 0)
        tot  = float(r.get("total",          0) or 0)
        t_l += qtd; t_r += tot
        for ci, v in enumerate([
            str(r.get("data", ""))[:10], dia_str, str(r.get("numero_ficha", "")),
            str(r.get("placa", "")), str(r.get("prefixo", "")), str(r.get("motorista", "")),
            str(r.get("fornecedor", "")), str(r.get("tipo_combustivel", ""))
        ]):
            ws.write(ri, ci, v, fd)
        ws.write(ri, 8,  qtd,  fn)
        ws.write(ri, 9,  vunt, fm)
        ws.write(ri, 10, tot,  fm)
        ws.write(ri, 11, str(r.get("horimetro",   "")), fd)
        ws.write(ri, 12, str(r.get("obra",         "")), fd)
        ws.write(ri, 13, str(r.get("observacao",   "")), fd)

    row_t = 5 + len(df)
    ws.merge_range(row_t, 0, row_t, 7, "TOTAL", ftx)
    ws.write(row_t, 8,  t_l, ft)
    ws.write(row_t, 9,  "",  ftx)
    ws.write(row_t, 10, t_r, ftm)
    ws.write(row_t, 11, "TOTAL", ftx)
    ws.write(row_t, 12, "", ftx)
    ws.write(row_t, 13, "", ftx)
    wb.close()
    buf.seek(0)
    return buf.getvalue()


def gerar_excel_tanque(df_ent: pd.DataFrame, df_sai: pd.DataFrame, df_transf: pd.DataFrame,
                       nome_tanque: str, periodo: str, obra: str) -> bytes:
    dias_pt = {0: "SEG", 1: "TER", 2: "QUA", 3: "QUI", 4: "SEX", 5: "SÁB", 6: "DOM"}
    import xlsxwriter
    buf = io.BytesIO()
    wb  = xlsxwriter.Workbook(buf)
    ws  = wb.add_worksheet("Tanque")
    fh  = wb.add_format({"bold": True, "font_size": 9, "border": 1, "align": "center", "bg_color": "#BDD7EE", "text_wrap": True})
    fd  = wb.add_format({"font_size": 9, "border": 1, "align": "center"})
    fn  = wb.add_format({"font_size": 9, "border": 1, "num_format": "#,##0.00", "align": "right"})
    fm  = wb.add_format({"font_size": 9, "border": 1, "num_format": '"R$" #,##0.00', "align": "right"})
    fok = wb.add_format({"font_size": 9, "border": 1, "num_format": "#,##0.00", "align": "right", "font_color": "#3B6D11"})
    flo = wb.add_format({"font_size": 9, "border": 1, "num_format": "#,##0.00", "align": "right", "font_color": "#854F0B"})
    ft  = wb.add_format({"bold": True, "font_size": 9, "border": 1, "num_format": "#,##0.00", "align": "right", "bg_color": "#D9D9D9"})
    ftx = wb.add_format({"bold": True, "font_size": 9, "border": 1, "align": "right", "bg_color": "#D9D9D9"})
    fti = wb.add_format({"bold": True, "font_size": 11, "align": "center", "bg_color": "#D9D9D9", "border": 1})
    flb = wb.add_format({"bold": True, "font_size": 9})
    fvl = wb.add_format({"font_size": 9})

    ws.write(0, 0, "OBRA:",   flb); ws.write(0, 2, obra.upper(),        fvl)
    ws.write(1, 0, "TANQUE:", flb); ws.write(1, 2, nome_tanque.upper(), fvl)
    ws.write(2, 0, "PERÍODO:",flb); ws.write(2, 2, periodo.upper(),     fvl)
    ws.merge_range(3, 0, 3, 15, f"CONTROLE DE TANQUE  —  {nome_tanque.upper()}", fti)

    heads  = ["DATA", "DIA", "TIPO", "FICHA", "PLACA", "PREF.", "VEÍ./OPERADOR",
              "PRODUTO / FORNEC.", "KM/HOR", "QTD ENTRADA (L)", "QTD SAÍDA (L)",
              "VL UNIT. (R$)", "TOTAL (R$)", "SALDO (L)", "OBRA", "OBS"]
    widths = [12, 7, 12, 10, 10, 8, 25, 22, 8, 14, 12, 12, 12, 12, 20, 20]
    ws.set_row(4, 30)
    for ci, (h, w) in enumerate(zip(heads, widths)):
        ws.write(4, ci, h, fh)
        ws.set_column(ci, ci, w)

    movs = []
    if not df_ent.empty:
        for _, r in df_ent.iterrows():
            movs.append({**r.to_dict(), "_tipo": "Entrada"})
    if not df_sai.empty:
        for _, r in df_sai.iterrows():
            movs.append({**r.to_dict(), "_tipo": "Saída"})
    if not df_transf.empty:
        for _, r in df_transf.iterrows():
            movs.append({**r.to_dict(), "_tipo": "Transf.Caminhão"})
    movs.sort(key=lambda x: str(x.get("data", "")))

    saldo = 0.0; t_ent = 0.0; t_sai = 0.0
    for ri, r in enumerate(movs, start=5):
        dia_str = ""
        try:
            dia_str = dias_pt[datetime.strptime(str(r.get("data", ""))[:10], "%Y-%m-%d").weekday()]
        except:
            pass
        qtd  = float(r.get("quantidade", 0) or 0)
        vunt = float(r.get("valor_unitario", 0) or 0)
        tot  = float(r.get("total", 0) or 0)
        tipo = r["_tipo"]
        if tipo == "Entrada":
            saldo += qtd; t_ent += qtd
        else:
            saldo -= qtd; t_sai += qtd
        q_ent = qtd if tipo == "Entrada" else 0
        q_sai = qtd if tipo != "Entrada" else 0
        forn_veic = r.get("fornecedor", "") if tipo == "Entrada" else r.get("motorista", r.get("caminhao_tanque", ""))
        prod_forn = r.get("combustivel", "") if tipo == "Entrada" else r.get("tipo_combustivel", r.get("produto", ""))
        ws.write(ri, 0,  str(r.get("data", ""))[:10], fd)
        ws.write(ri, 1,  dia_str, fd)
        ws.write(ri, 2,  tipo, fd)
        ws.write(ri, 3,  str(r.get("numero_ficha", "")), fd)
        ws.write(ri, 4,  str(r.get("placa", "")), fd)
        ws.write(ri, 5,  str(r.get("prefixo", "")), fd)
        ws.write(ri, 6,  forn_veic, fd)
        ws.write(ri, 7,  prod_forn, fd)
        ws.write(ri, 8,  str(r.get("horimetro", "")), fd)
        ws.write(ri, 9,  q_ent if q_ent else "", fn if q_ent else fd)
        ws.write(ri, 10, q_sai if q_sai else "", fn if q_sai else fd)
        ws.write(ri, 11, vunt, fm)
        ws.write(ri, 12, tot,  fm)
        ws.write(ri, 13, saldo, fok if saldo >= 500 else flo)
        ws.write(ri, 14, str(r.get("obra", "")), fd)
        ws.write(ri, 15, str(r.get("observacao", "")), fd)

    row_t = 5 + len(movs)
    ws.write(row_t, 0, "TOTAL", ftx)
    ws.merge_range(row_t, 1, row_t, 8, "", ftx)
    ws.write(row_t, 9,  t_ent, ft)
    ws.write(row_t, 10, t_sai, ft)
    ws.write(row_t, 11, "", ftx); ws.write(row_t, 12, "", ftx)
    ws.write(row_t, 13, t_ent - t_sai, ft)
    ws.write(row_t, 14, "", ftx); ws.write(row_t, 15, "", ftx)
    wb.close()
    buf.seek(0)
    return buf.getvalue()


def gerar_excel_limpo(df: pd.DataFrame, nome_aba: str = "Relatório") -> bytes:
    df = df.fillna("").copy()
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as w:
        df.to_excel(w, index=False, sheet_name=nome_aba)
        ws = w.sheets[nome_aba]
        for i, col in enumerate(df.columns):
            try:
                sz = max(len(str(col)), df[col].astype(str).str.len().max())
                ws.set_column(i, i, min(int(sz) + 2, 50))
            except:
                ws.set_column(i, i, 15)
    return buf.getvalue()


def gerar_pdf(df: pd.DataFrame, tipo: str, titulo_esq: str, sub_esq: str, per_esq: str,
              dados_dir: dict, titulo_tab: str) -> bytes:
    df = df.fillna("").copy()
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.rect(10, 10, 110, 22)
    x = 12
    if os.path.exists("logo.png"):
        try:
            pdf.image("logo.png", x=12, y=12, h=18)
            x = 48
        except:
            pass
    pdf.set_xy(x, 12); pdf.set_font("Arial", "B", 9)
    pdf.cell(0, 6, titulo_esq.upper(), ln=1)
    pdf.set_x(x); pdf.cell(0, 6, sub_esq.upper(), ln=1)
    pdf.set_x(x); pdf.cell(0, 6, per_esq.upper(), ln=1)
    pdf.rect(125, 10, 162, 22)
    pdf.set_xy(127, 12); pdf.set_font("Arial", "B", 9)
    for i, (k, v) in enumerate(dados_dir.items()):
        if i % 2 == 0:
            pdf.set_x(127); pdf.cell(80, 5, f"{k}: {v}")
        else:
            pdf.cell(80, 5, f"{k}: {v}", ln=1)
    pdf.set_y(35); pdf.set_font("Arial", "B", 10)
    pdf.cell(277, 8, titulo_tab.upper(), border=1, align="C", ln=1)
    pdf.set_font("Arial", "B", 7); pdf.set_fill_color(220, 220, 220)

    if tipo == "SAIDAS":
        cols = [("DATA", 16), ("FICHA", 15), ("PLACA", 14), ("PREF.", 12),
                ("MÁQUINA / MOTORISTA", 45), ("PRODUTO", 18), ("QTD (L)", 14),
                ("V.UNIT.", 14), ("TOTAL (R$)", 20), ("KM/HOR", 13), ("OBRA", 22), ("OBS", 47)]
    elif tipo == "TANQUE":
        cols = [("DATA", 14), ("TIPO", 14), ("FICHA", 14), ("PLACA", 13), ("PREF.", 10),
                ("OPERADOR / FORNEC.", 36), ("PRODUTO", 16), ("KM/H", 10),
                ("ENTRADA(L)", 17), ("SAÍDA(L)", 14), ("V.UNIT.", 13),
                ("TOTAL", 15), ("SALDO(L)", 14), ("OBRA", 20), ("OBS", 17)]
    else:
        cols = [("DATA", 18), ("NF/FICHA", 22), ("DISTRIBUIDORA", 55), ("TANQUE", 38),
                ("PRODUTO", 22), ("QTD (L)", 20), ("V.UNIT.", 20), ("TOTAL (R$)", 23),
                ("OBRA", 30), ("OBS", 29)]

    for n, w in cols:
        pdf.cell(w, 7, n, border=1, align="C", fill=True)
    pdf.ln(); pdf.set_font("Arial", "", 7)
    t_l = 0.0; t_r = 0.0; saldo = 0.0

    for _, r in df.iterrows():
        q  = float(r.get("quantidade",    0) or 0)
        v  = float(r.get("valor_unitario",0) or 0)
        t  = float(r.get("total",         0) or 0)
        qe = float(r.get("qtd_entrada",   0) or 0)
        qs = float(r.get("qtd_saida",     0) or 0)
        saldo += qe - qs
        if tipo == "SAIDAS":
            for val, w in [
                (str(r.get("data", ""))[:10], 16), (str(r.get("numero_ficha", ""))[:14], 15),
                (str(r.get("placa", ""))[:8], 14), (str(r.get("prefixo", ""))[:8], 12),
                (str(r.get("motorista", ""))[:30], 45), (str(r.get("tipo_combustivel", ""))[:12], 18)
            ]:
                pdf.cell(w, 6, val, border=1, align="C")
            pdf.cell(14, 6, f"{q:,.2f}",   border=1, align="R")
            pdf.cell(14, 6, f"{v:,.2f}",   border=1, align="R")
            pdf.cell(20, 6, f"R${t:,.2f}", border=1, align="R")
            pdf.cell(13, 6, str(r.get("horimetro", ""))[:7], border=1, align="C")
            pdf.cell(22, 6, str(r.get("obra", ""))[:14],     border=1, align="C")
            pdf.cell(47, 6, str(r.get("observacao", ""))[:30], border=1, align="L")
            t_l += q; t_r += t
        elif tipo == "TANQUE":
            pdf.cell(14, 6, str(r.get("data", ""))[:10],         border=1, align="C")
            pdf.cell(14, 6, str(r.get("tipo", ""))[:10],         border=1, align="C")
            pdf.cell(14, 6, str(r.get("numero_ficha", ""))[:12], border=1, align="C")
            pdf.cell(13, 6, str(r.get("placa", ""))[:7],         border=1, align="C")
            pdf.cell(10, 6, str(r.get("prefixo", ""))[:7],       border=1, align="C")
            pdf.cell(36, 6, str(r.get("motorista_forn", ""))[:23], border=1, align="L")
            pdf.cell(16, 6, str(r.get("produto", ""))[:10],      border=1, align="C")
            pdf.cell(10, 6, str(r.get("horimetro", ""))[:7],     border=1, align="C")
            pdf.cell(17, 6, f"{qe:,.1f}" if qe else "-",         border=1, align="R")
            pdf.cell(14, 6, f"{qs:,.1f}" if qs else "-",         border=1, align="R")
            pdf.cell(13, 6, f"{v:,.2f}",                         border=1, align="R")
            pdf.cell(15, 6, f"R${t:,.2f}",                       border=1, align="R")
            pdf.cell(14, 6, f"{saldo:,.1f}",                     border=1, align="R")
            pdf.cell(20, 6, str(r.get("obra", ""))[:13],         border=1, align="C")
            pdf.cell(17, 6, str(r.get("observacao", ""))[:11],   border=1, align="L")
            t_l += (qe or qs); t_r += t
        else:
            pdf.cell(18, 6, str(r.get("data", ""))[:10],         border=1, align="C")
            pdf.cell(22, 6, str(r.get("numero_ficha", ""))[:14], border=1, align="C")
            pdf.cell(55, 6, str(r.get("fornecedor", ""))[:35],   border=1, align="L")
            pdf.cell(38, 6, str(r.get("nome_tanque", ""))[:22],  border=1, align="C")
            pdf.cell(22, 6, str(r.get("combustivel", ""))[:12],  border=1, align="C")
            pdf.cell(20, 6, f"{q:,.2f}",                         border=1, align="R")
            pdf.cell(20, 6, f"{v:,.2f}",                         border=1, align="R")
            pdf.cell(23, 6, f"R${t:,.2f}",                       border=1, align="R")
            pdf.cell(30, 6, str(r.get("obra", ""))[:18],         border=1, align="C")
            pdf.cell(29, 6, str(r.get("observacao", ""))[:18],   border=1, align="L")
            t_l += q; t_r += t
        pdf.ln()

    pdf.set_font("Arial", "B", 8)
    if tipo == "SAIDAS":
        pdf.cell(136, 8, "TOTAIS GERAIS", border=1, align="R")
        pdf.cell(14,  8, f"{t_l:,.2f}",  border=1, align="R")
        pdf.cell(14,  8, "-",            border=1, align="C")
        pdf.cell(20,  8, f"R$ {t_r:,.2f}", border=1, align="R")
        pdf.cell(93,  8, "",             border=1)
    elif tipo == "TANQUE":
        pdf.cell(138, 8, "TOTAIS",       border=1, align="R")
        pdf.cell(17,  8, f"{t_l:,.2f}", border=1, align="R")
        pdf.cell(14,  8, "-",           border=1, align="C")
        pdf.cell(13,  8, "-",           border=1, align="C")
        pdf.cell(15,  8, f"R$ {t_r:,.2f}", border=1, align="R")
        pdf.cell(14,  8, f"{saldo:,.1f}", border=1, align="R")
        pdf.cell(54,  8, "",            border=1)
    else:
        pdf.cell(167, 8, "TOTAIS GERAIS", border=1, align="R")
        pdf.cell(20,  8, f"{t_l:,.2f}",  border=1, align="R")
        pdf.cell(20,  8, "-",            border=1, align="C")
        pdf.cell(23,  8, f"R$ {t_r:,.2f}", border=1, align="R")
        pdf.cell(59,  8, "",             border=1)

    return pdf.output(dest="S").encode("latin-1")


# ═══════════════════════════════════════════════════════════════════
# LOGIN
# ═══════════════════════════════════════════════════════════════════
for k, v in [("logged_in", False), ("usuario_logado", ""), ("perfil_logado", "")]:
    if k not in st.session_state:
        st.session_state[k] = v

if not st.session_state.logged_in:
    st.markdown("""
    <style>
    [data-testid="stAppViewContainer"] {
        background: linear-gradient(rgba(15, 25, 35, 0.7), rgba(15, 25, 35, 0.7)), 
        url('https://images.unsplash.com/photo-1463171379579-3fdfb86d6285?q=80&w=2070') no-repeat center center fixed !important;
        background-size: cover !important;
    }
    [data-testid="stHeader"] { background: transparent !important; }
    [data-testid="stSidebar"] { display: none; }
    div[data-testid="stForm"] { margin-top: -20px; }
    img { margin-bottom: -15px; }
    </style>
    """, unsafe_allow_html=True)

    st.write("<br>", unsafe_allow_html=True)
    c1, c2, c3 = st.columns([1, 2.5, 1])
    with c2:
        with st.form("login"):
            if os.path.exists("logo.png"):
                col_l1, col_l2, col_l3 = st.columns([1, 2, 1])
                with col_l2:
                    st.image("logo.png", width=280)
                    st.markdown("<div style='margin-top:-25px'></div>", unsafe_allow_html=True)

            st.markdown("<h2 style='text-align:center; color:#1E293B; font-weight:700; margin-top:0; margin-bottom:0.3rem;'>Acesso Restrito</h2>", unsafe_allow_html=True)
            u = st.text_input("Usuário")
            p = st.text_input("Senha", type="password")
            st.write("<br>", unsafe_allow_html=True)

            if st.form_submit_button("ENTRAR NO SISTEMA", use_container_width=True):
                try:
                    res = supabase.table("usuarios").select("*").eq("login", u).eq("senha", p).execute()
                    if res.data:
                        st.session_state.logged_in       = True
                        st.session_state.usuario_logado  = res.data[0]["nome"]
                        st.session_state.perfil_logado   = res.data[0]["perfil"]
                        st.rerun()
                    else:
                        st.error("❌ Usuário ou senha incorretos.")
                except:
                    if u == st.secrets.get("ADMIN_USER", "admin") and p == st.secrets.get("ADMIN_PASS", "obra2026"):
                        st.session_state.logged_in       = True
                        st.session_state.usuario_logado  = "Admin"
                        st.session_state.perfil_logado   = "Admin"
                        st.rerun()
                    else:
                        st.error("❌ Usuário ou senha incorretos.")
    st.stop()

# ═══════════════════════════════════════════════════════════════════
# SIDEBAR / MENU
# ═══════════════════════════════════════════════════════════════════
with st.sidebar:
    if os.path.exists("logo.png"):
        c1, c2, c3 = st.columns([1, 2, 1])
        with c2:
            st.image("logo.png", width=300)

    st.markdown(
        f"<div style='text-align:center;color:#1D9E75;font-size:13px;font-weight:bold;'>👤 {st.session_state.usuario_logado}</div>",
        unsafe_allow_html=True
    )
    st.divider()

    opcoes = [
        "🏠 Painel Início",
        "⛽ Lançar Abastecimento",
        "🔄 Transferência Caminhão-Tanque",
        "🛢️ Tanques / Estoque",
        "🚚 Boletim de Transporte",
        "🚜 Frota e Equipamentos",
        "🏗️ Obras",
        "🏪 Fornecedores",
        "📋 Relatórios e Fechamentos",
    ]
    if st.session_state.perfil_logado == "Admin":
        opcoes.append("👥 Usuários e Acessos")

    menu = st.radio("", opcoes, label_visibility="collapsed")
    st.divider()
    st.markdown("<br><br><br><br><br>", unsafe_allow_html=True)
    if st.button("Sair"):
        st.session_state.logged_in      = False
        st.session_state.usuario_logado = ""
        st.session_state.perfil_logado  = ""
        st.rerun()
    st.caption("☁️ Supabase — Tempo Real")

# ════════════════════════════════════════════════════════════════════
# 1 · PAINEL INÍCIO
# ════════════════════════════════════════════════════════════════════
if menu == "🏠 Painel Início":
    st.markdown("## 🏠 Centro de Comando")

    with st.spinner("Carregando dados..."):
        df_tanq = get_data("tanques")
        df_ab   = get_data("abastecimentos")
        df_prod = get_data("producao")

        if not df_ab.empty and "status" in df_ab.columns:
            df_ab = df_ab[df_ab["status"] == "ATIVO"]

    # ── TANQUES ──────────────────────────────────────────────────────
    if not df_tanq.empty:
        st.subheader("🛢️ Situação dos Tanques / Comboios")
        cols_t = st.columns(min(len(df_tanq), 5))
        saldos = {row["nome"]: calcular_saldo(row["nome"]) for _, row in df_tanq.iterrows()}

        for idx, row in df_tanq.iterrows():
            nm  = row["nome"]
            cap = float(row.get("capacidade", 0) or 0)
            sd  = saldos.get(nm, 0)
            lim = cap * 0.15 if cap > 0 else 500
            low = sd <= lim
            cls = "banner-low" if low else "banner-ok"
            ic  = "⚠️" if low else "✅"
            pct = f" / {sd / cap * 100:.0f}%" if cap > 0 else ""

            with cols_t[idx % len(cols_t)]:
                st.markdown(f"<div class='{cls}'>{ic} <strong>{nm}</strong><br>{sd:,.1f} L{pct}</div>", unsafe_allow_html=True)
                if cap > 0:
                    st.progress(max(0.0, min(sd / cap, 1.0)))

    # ── FILTROS ───────────────────────────────────────────────────────
    st.markdown("#### 📅 Indicadores do Período")
    fc1, fc2, fc3 = st.columns(3)
    d_ini = fc1.date_input("De", value=date.today().replace(day=1))
    d_fim = fc2.date_input("Até", value=date.today())

    obras_painel = lista_obras(incluir_todas=True)
    obra_filtro  = fc3.selectbox("Obra", obras_painel) if obras_painel else "TODAS"

    t_gasto = 0; t_litros = 0; t_carradas = 0; t_ton = 0; t_ton_cbuq = 0

    if not df_ab.empty and "data" in df_ab.columns:
        df_ab["data_dt"] = pd.to_datetime(df_ab["data"], errors="coerce")
        daf = df_ab[
            (df_ab["data_dt"].notna()) &
            (df_ab["data_dt"].dt.date >= d_ini) &
            (df_ab["data_dt"].dt.date <= d_fim)
        ]
        if obra_filtro and obra_filtro != "TODAS" and "obra" in daf.columns:
            daf = daf[daf["obra"] == obra_filtro]
        if not daf.empty:
            t_gasto  = pd.to_numeric(daf.get("total",      0), errors="coerce").sum()
            t_litros = pd.to_numeric(daf.get("quantidade", 0), errors="coerce").sum()

    if not df_prod.empty and "data" in df_prod.columns:
        df_prod["data_dt"] = pd.to_datetime(df_prod["data"], errors="coerce")
        dpf = df_prod[
            (df_prod["data_dt"].notna()) &
            (df_prod["data_dt"].dt.date >= d_ini) &
            (df_prod["data_dt"].dt.date <= d_fim)
        ]
        if obra_filtro and obra_filtro != "TODAS" and "obra" in dpf.columns:
            dpf = dpf[dpf["obra"] == obra_filtro]
        if not dpf.empty:
            t_carradas = pd.to_numeric(dpf.get("carradas",  0), errors="coerce").sum()
            t_ton      = pd.to_numeric(dpf.get("toneladas", 0), errors="coerce").sum()
            dc = dpf[dpf.get("tipo_operacao", pd.Series(dtype=str)).isin(["Transporte de Massa/CBUQ", "Venda de Massa"])]
            if not dc.empty:
                t_ton_cbuq = pd.to_numeric(dc.get("toneladas", 0), errors="coerce").sum()

    # ── KPIs ──────────────────────────────────────────────────────────
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("💰 Gasto Combustível", f"R$ {t_gasto:,.2f}")
    c2.metric("⛽ Litros",            f"{t_litros:,.1f} L")
    c3.metric("🏗️ Ton CBUQ",         f"{t_ton_cbuq:,.1f} t")
    c4.metric("🚚 Viagens",           int(t_carradas))

    st.divider()
    st.markdown("#### ⚙️ KPIs de Eficiência")
    c5, c6, c7 = st.columns(3)
    custo_ton   = t_gasto  / t_ton_cbuq  if t_ton_cbuq  > 0 else 0
    litros_ton  = t_litros / t_ton_cbuq  if t_ton_cbuq  > 0 else 0
    litros_vg   = t_litros / t_carradas  if t_carradas  > 0 else 0
    c5.metric("Custo / Ton CBUQ", f"R$ {custo_ton:,.2f}")
    c6.metric("Litros / Ton",     f"{litros_ton:,.2f} L")
    c7.metric("Litros / Viagem",  f"{litros_vg:,.1f} L")

    # ── GRÁFICOS ──────────────────────────────────────────────────────
    if not df_ab.empty:
        df_filtrado = df_ab[
            (df_ab["data_dt"].notna()) &
            (df_ab["data_dt"].dt.date >= d_ini) &
            (df_ab["data_dt"].dt.date <= d_fim)
        ]
        if obra_filtro and obra_filtro != "TODAS" and "obra" in df_filtrado.columns:
            df_filtrado = df_filtrado[df_filtrado["obra"] == obra_filtro]

        df_filtrado["Mês"]     = df_filtrado["data_dt"].dt.strftime("%m/%Y")
        df_filtrado["total_n"] = pd.to_numeric(df_filtrado.get("total", 0), errors="coerce").fillna(0)

        g = df_filtrado.groupby("Mês")["total_n"].sum().reset_index()

        col_g1, col_g2 = st.columns(2)

        with col_g1:
            if not g.empty:
                st.subheader("📊 Gastos por Mês")
                fig = px.bar(
                    g, x="Mês", y="total_n", text="total_n",
                    labels={"total_n": "Total Gasto (R$)", "Mês": "Mês/Ano"},
                    color_discrete_sequence=["#0A58CA"]
                )
                fig.update_traces(texttemplate="R$ %{text:,.2f}", textposition="outside", marker_line_width=0)
                fig.update_layout(
                    plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                    margin=dict(t=20, b=20, l=0, r=0),
                    xaxis=dict(showgrid=False),
                    yaxis=dict(showgrid=True, gridcolor="#E2E8F0", tickformat=",.2f")
                )
                st.plotly_chart(fig, use_container_width=True)

        with col_g2:
            # Gráfico de consumo por veículo
            if "prefixo" in df_filtrado.columns:
                df_filtrado["qtd_n"] = pd.to_numeric(df_filtrado.get("quantidade", 0), errors="coerce").fillna(0)
                gv = df_filtrado.groupby("prefixo")["qtd_n"].sum().reset_index().sort_values("qtd_n", ascending=False).head(10)
                if not gv.empty:
                    st.subheader("🚜 Top 10 Veículos por Consumo (L)")
                    fig2 = px.bar(
                        gv, x="qtd_n", y="prefixo", orientation="h",
                        labels={"qtd_n": "Litros", "prefixo": "Veículo"},
                        color_discrete_sequence=["#1D9E75"]
                    )
                    fig2.update_layout(
                        plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                        margin=dict(t=20, b=20, l=0, r=0),
                        yaxis=dict(autorange="reversed")
                    )
                    st.plotly_chart(fig2, use_container_width=True)

        # Gráfico de consumo por obra
        if "obra" in df_filtrado.columns and not df_filtrado["obra"].replace("", pd.NA).dropna().empty:
            df_filtrado["total_n2"] = pd.to_numeric(df_filtrado.get("total", 0), errors="coerce").fillna(0)
            go_df = df_filtrado[df_filtrado["obra"].notna() & (df_filtrado["obra"] != "")].groupby("obra")["total_n2"].sum().reset_index()
            if not go_df.empty:
                st.subheader("🏗️ Gastos por Obra")
                fig3 = px.pie(go_df, values="total_n2", names="obra",
                              color_discrete_sequence=px.colors.qualitative.Set2)
                fig3.update_layout(margin=dict(t=20, b=20, l=0, r=0))
                st.plotly_chart(fig3, use_container_width=True)
    else:
        st.info("ℹ️ Nenhum abastecimento registrado para o período selecionado.")


# ════════════════════════════════════════════════════════════════════
# 2 · LANÇAR ABASTECIMENTO
# ════════════════════════════════════════════════════════════════════
elif menu == "⛽ Lançar Abastecimento":
    st.markdown("## ⛽ Lançar Aquisição de Combustível")

    df_v = get_data("veiculos")
    df_f = get_data("fornecedores")
    df_a = get_data("abastecimentos")
    df_t = get_data("tanques")

    if df_v.empty:
        st.warning("⚠️ Cadastre veículos primeiro.")
        st.stop()

    # Proteção de colunas
    for col in ["categoria", "tipo_veiculo", "tipo_combustivel_padrao", "motorista", "placa"]:
        if col not in df_v.columns:
            df_v[col] = ""

    # Exclui caminhões-tanque do seletor (eles têm aba própria)
    df_v_normal = df_v[df_v["tipo_veiculo"] != "Caminhão-Tanque"] if "tipo_veiculo" in df_v.columns else df_v

    v_sel     = st.selectbox("🚜 Máquina / Veículo", df_v_normal["prefixo"].tolist())
    info_v    = df_v_normal[df_v_normal["prefixo"] == v_sel].iloc[0]

    comb_padrao     = info_v.get("tipo_combustivel_padrao", "Diesel S10")
    motorista_padrao = info_v.get("motorista", "")
    placa_padrao    = info_v.get("placa", "")

    m_hor = 0
    if not df_a.empty and "horimetro" in df_a.columns:
        hist = df_a[df_a["prefixo"] == v_sel]
        if not hist.empty:
            m_hor = float(pd.to_numeric(hist["horimetro"], errors="coerce").max() or 0)

    st.info(f"⛽ {comb_padrao} | 🪪 {placa_padrao} | ⏱️ Último KM/Hor: {m_hor:,.1f}")

    origem = st.radio("Origem do Combustível:", ["Posto Externo", "Tanque Interno"], horizontal=True)

    # Mostra saldo do tanque em tempo real ao selecionar Tanque Interno
    saldo_tanque_atual = None
    if origem == "Tanque Interno" and not df_t.empty:
        n_tanq_preview = st.selectbox("Tanque (pré-visualização de saldo)", df_t["nome"].tolist(), key="prev_tanq")
        saldo_tanque_atual = calcular_saldo(n_tanq_preview)
        cor = "banner-ok" if saldo_tanque_atual >= 500 else "banner-low"
        st.markdown(f"<div class='{cor}'>🛢️ Saldo atual de <strong>{n_tanq_preview}</strong>: <strong>{saldo_tanque_atual:,.1f} L</strong></div>", unsafe_allow_html=True)

    obras_lista = lista_obras()

    with st.form("form_ab", clear_on_submit=True):
        c1, c2, c3 = st.columns(3)
        data_ab  = c1.date_input("Data")
        ficha    = c2.text_input("Ficha / Nota Fiscal")
        motorista = c3.text_input("Motorista", value=motorista_padrao)

        c4, c5, c6 = st.columns(3)
        if origem == "Posto Externo":
            posto  = c4.selectbox("Fornecedor", df_f["nome"].tolist() if not df_f.empty else ["Sem cadastro"])
            n_tanq = None
        else:
            n_tanq = c4.selectbox("Tanque", df_t["nome"].tolist() if not df_t.empty else [], key="tanq_form")
            posto  = "Estoque Próprio"

        hor = c5.number_input("KM / Horímetro", value=m_hor)
        obs = c6.text_input("Observação")

        c7, c8, c9 = st.columns(3)
        litros = c7.number_input("Litros", min_value=0.0)
        preco  = c8.number_input("Preço (R$/L)", min_value=0.0)

        # Campo de obra — vincula o abastecimento à obra
        if obras_lista:
            obra_ab = c9.selectbox("Obra / Projeto", obras_lista)
        else:
            obra_ab = c9.text_input("Obra / Projeto")

        total = litros * preco
        st.info(f"💰 Total calculado: R$ {total:,.2f}")

        # Alerta se litros > saldo do tanque
        if origem == "Tanque Interno" and saldo_tanque_atual is not None and litros > saldo_tanque_atual:
            st.warning(f"⚠️ Quantidade ({litros:,.1f} L) excede o saldo do tanque ({saldo_tanque_atual:,.1f} L)!")

        if st.form_submit_button("💾 Salvar Abastecimento", use_container_width=True):
            if litros <= 0:
                st.error("⚠️ Quantidade de litros inválida.")
            else:
                dados = {
                    "data":             str(data_ab),
                    "numero_ficha":     ficha,
                    "origem":           origem,
                    "nome_tanque":      n_tanq,
                    "prefixo":          v_sel,
                    "placa":            placa_padrao,
                    "motorista":        motorista.upper(),
                    "tipo_combustivel": comb_padrao,
                    "quantidade":       litros,
                    "valor_unitario":   preco,
                    "total":            total,
                    "fornecedor":       posto,
                    "horimetro":        hor,
                    "obra":             obra_ab,
                    "observacao":       obs,
                    "status":           "ATIVO",
                    "criado_por":       st.session_state.usuario_logado,
                }
                ok = insert_data("abastecimentos", dados)
                if ok:
                    st.success("✅ Abastecimento salvo com sucesso!")
                    time.sleep(1)
                    st.rerun()

    # ── LISTAGEM COM FILTROS ──────────────────────────────────────────
    st.divider()
    st.subheader("📋 Abastecimentos Registrados")

    if df_a.empty:
        st.info("Nenhum registro encontrado.")
    else:
        df_a = df_a.sort_values("data", ascending=False).fillna("")

        # Filtros de listagem
        fl1, fl2, fl3, fl4 = st.columns(4)
        f_di   = fl1.date_input("De",    value=date.today().replace(day=1), key="f_di_ab")
        f_df   = fl2.date_input("Até",   value=date.today(),                key="f_df_ab")
        f_veic = fl3.selectbox("Veículo", ["TODOS"] + df_v["prefixo"].tolist(), key="f_veic_ab")
        obras_f = lista_obras(incluir_todas=True)
        f_obra = fl4.selectbox("Obra",   obras_f if obras_f else ["TODAS"],  key="f_obra_ab")

        df_a["data_dt"] = pd.to_datetime(df_a["data"], errors="coerce").dt.date
        df_a_fil = df_a[(df_a["data_dt"] >= f_di) & (df_a["data_dt"] <= f_df)]
        if f_veic != "TODOS":
            df_a_fil = df_a_fil[df_a_fil["prefixo"] == f_veic]
        if f_obra not in ("TODAS", "") and "obra" in df_a_fil.columns:
            df_a_fil = df_a_fil[df_a_fil["obra"] == f_obra]

        ativos     = df_a_fil[df_a_fil["status"] == "ATIVO"]
        cancelados = df_a_fil[df_a_fil["status"] != "ATIVO"]

        tab1, tab2 = st.tabs([f"✅ Ativos ({len(ativos)})", f"❌ Cancelados ({len(cancelados)})"])

        with tab1:
            if ativos.empty:
                st.info("Nenhum registro ativo no período/filtro.")
            else:
                for r in ativos.head(50).to_dict("records"):
                    obra_tag = f" | 🏗️ {r.get('obra', '')}" if r.get("obra") else ""
                    with st.expander(
                        f"📅 {r.get('data', '')} | 🚜 {r.get('prefixo', '')} | "
                        f"⛽ {r.get('quantidade', 0)} L | 💰 R$ {float(r.get('total', 0) or 0):,.2f}{obra_tag}"
                    ):
                        ec1, ec2, ec3, ec4 = st.columns(4)
                        ec1.write(f"**Motorista:** {r.get('motorista', '')}")
                        ec2.write(f"**Placa:** {r.get('placa', '')}")
                        ec3.write(f"**Fornecedor:** {r.get('fornecedor', '')}")
                        ec4.write(f"**Origem:** {r.get('origem', '')}")
                        ec5, ec6, ec7, ec8 = st.columns(4)
                        ec5.write(f"**Ficha:** {r.get('numero_ficha', '')}")
                        ec6.write(f"**KM/Hor:** {r.get('horimetro', '')}")
                        ec7.write(f"**Obra:** {r.get('obra', '-')}")
                        ec8.write(f"**Obs:** {r.get('observacao', '')}")

                        # ── EDIÇÃO INLINE ──────────────────────────────
                        with st.form(f"edit_form_{r.get('id')}"):
                            st.markdown("**✏️ Editar Registro**")
                            ne1, ne2, ne3 = st.columns(3)
                            new_litros = ne1.number_input("Litros", value=float(r.get("quantidade", 0) or 0), min_value=0.0, key=f"nl_{r.get('id')}")
                            new_preco  = ne2.number_input("Preço (R$/L)", value=float(r.get("valor_unitario", 0) or 0), min_value=0.0, key=f"np_{r.get('id')}")
                            obras_edit = lista_obras()
                            obra_atual = r.get("obra", "")
                            if obras_edit:
                                idx_obra = obras_edit.index(obra_atual) if obra_atual in obras_edit else 0
                                new_obra = ne3.selectbox("Obra", obras_edit, index=idx_obra, key=f"no_{r.get('id')}")
                            else:
                                new_obra = ne3.text_input("Obra", value=obra_atual, key=f"no_{r.get('id')}")
                            ne4, ne5 = st.columns(2)
                            new_hor  = ne4.number_input("KM/Hor", value=float(r.get("horimetro", 0) or 0), min_value=0.0, key=f"nh_{r.get('id')}")
                            new_obs  = ne5.text_input("Observação", value=r.get("observacao", ""), key=f"nobs_{r.get('id')}")
                            col_save, col_cancel = st.columns(2)
                            if col_save.form_submit_button("💾 Salvar Edição", use_container_width=True):
                                ok_edit = update_data("abastecimentos", r.get("id"), {
                                    "quantidade":     new_litros,
                                    "valor_unitario": new_preco,
                                    "total":          new_litros * new_preco,
                                    "horimetro":      new_hor,
                                    "obra":           new_obra,
                                    "observacao":     new_obs,
                                })
                                if ok_edit:
                                    st.success("✅ Registro atualizado!")
                                    time.sleep(1); st.rerun()
                            if col_cancel.form_submit_button("❌ Cancelar Registro", use_container_width=True):
                                supabase.table("abastecimentos").update({"status": "CANCELADO"}).eq("id", r.get("id")).execute()
                                st.warning("Registro cancelado.")
                                time.sleep(1); st.rerun()

        with tab2:
            if cancelados.empty:
                st.info("Nenhum registro cancelado no período/filtro.")
            else:
                for r in cancelados.head(50).to_dict("records"):
                    c1, c2 = st.columns([5, 1])
                    c1.write(f"❌ {r.get('data', '')} | {r.get('prefixo', '')} | {r.get('quantidade', 0)} L | Obra: {r.get('obra', '-')}")
                    if c2.button("↩️ Restaurar", key=f"restore_{r.get('id')}"):
                        supabase.table("abastecimentos").update({"status": "ATIVO"}).eq("id", r.get("id")).execute()
                        st.success("Restaurado!")
                        time.sleep(1); st.rerun()


# ════════════════════════════════════════════════════════════════════
# 2B · TRANSFERÊNCIA CAMINHÃO-TANQUE
# ════════════════════════════════════════════════════════════════════
elif menu == "🔄 Transferência Caminhão-Tanque":
    st.markdown("## 🔄 Transferência de Combustível — Caminhão-Tanque")
    st.markdown("""
    <div class='banner-info'>
    ℹ️ Use esta aba para registrar quando um <strong>caminhão-tanque</strong> retira combustível
    do tanque fixo e abastece veículos em campo. Cada transferência fica rastreada com a obra atendida.
    </div>
    """, unsafe_allow_html=True)

    df_v    = get_data("veiculos")
    df_t    = get_data("tanques")
    df_tr   = get_data("transferencias_tanque")

    # Filtra apenas caminhões-tanque cadastrados
    df_ct = pd.DataFrame()
    if not df_v.empty and "tipo_veiculo" in df_v.columns:
        df_ct = df_v[df_v["tipo_veiculo"] == "Caminhão-Tanque"]

    obras_lista = lista_obras()

    tab_reg, tab_hist = st.tabs(["➕ Registrar Transferência", "📋 Histórico de Transferências"])

    with tab_reg:
        if df_t.empty:
            st.warning("⚠️ Cadastre pelo menos um tanque fixo antes de registrar transferências.")
        else:
            # Saldo dos tanques disponíveis
            st.markdown("#### 🛢️ Saldo dos Tanques Disponíveis")
            cols_sd = st.columns(min(len(df_t), 4))
            for idx_t, row_t in df_t.iterrows():
                sd = calcular_saldo(row_t["nome"])
                cor = "banner-ok" if sd >= 500 else "banner-low"
                with cols_sd[idx_t % min(len(df_t), 4)]:
                    st.markdown(f"<div class='{cor}'><strong>{row_t['nome']}</strong><br>{sd:,.1f} L</div>", unsafe_allow_html=True)

            st.divider()

            with st.form("form_transf", clear_on_submit=True):
                st.markdown("#### 📋 Dados da Transferência")
                ct1, ct2, ct3 = st.columns(3)
                data_tr  = ct1.date_input("Data da Transferência")
                ficha_tr = ct2.text_input("Ficha / Documento")
                tanq_orig = ct3.selectbox("Tanque de Origem (Fixo)", df_t["nome"].tolist())

                ct4, ct5, ct6 = st.columns(3)
                if not df_ct.empty:
                    caminhao_sel = ct4.selectbox("Caminhão-Tanque", df_ct["prefixo"].tolist())
                    info_ct = df_ct[df_ct["prefixo"] == caminhao_sel].iloc[0]
                    motorista_ct = info_ct.get("motorista", "")
                    placa_ct     = info_ct.get("placa", "")
                else:
                    caminhao_sel = ct4.text_input("Caminhão-Tanque (prefixo)")
                    motorista_ct = ""
                    placa_ct     = ""

                motorista_tr = ct5.text_input("Motorista", value=motorista_ct)
                placa_tr     = ct6.text_input("Placa", value=placa_ct)

                ct7, ct8, ct9 = st.columns(3)
                qtd_tr   = ct7.number_input("Quantidade Transferida (L)", min_value=0.0)
                vunt_tr  = ct8.number_input("Valor Unitário (R$/L)", min_value=0.0)
                produto_tr = ct9.selectbox("Produto", ["Diesel S10", "Diesel S500", "Gasolina Comum"])

                ct10, ct11 = st.columns(2)
                if obras_lista:
                    obra_tr = ct10.selectbox("Obra Atendida", obras_lista)
                else:
                    obra_tr = ct10.text_input("Obra Atendida")
                obs_tr = ct11.text_input("Observação")

                total_tr = qtd_tr * vunt_tr
                st.info(f"💰 Total: R$ {total_tr:,.2f}")

                # Alerta de saldo
                saldo_orig = calcular_saldo(tanq_orig)
                if qtd_tr > saldo_orig:
                    st.warning(f"⚠️ Quantidade ({qtd_tr:,.1f} L) excede o saldo do tanque {tanq_orig} ({saldo_orig:,.1f} L)!")

                if st.form_submit_button("💾 Registrar Transferência", use_container_width=True):
                    if qtd_tr <= 0:
                        st.error("⚠️ Quantidade inválida.")
                    else:
                        dados_tr = {
                            "data":             str(data_tr),
                            "numero_ficha":     ficha_tr,
                            "tanque_origem":    tanq_orig,
                            "caminhao_tanque":  caminhao_sel,
                            "placa":            placa_tr,
                            "motorista":        motorista_tr.upper(),
                            "produto":          produto_tr,
                            "quantidade":       qtd_tr,
                            "valor_unitario":   vunt_tr,
                            "total":            total_tr,
                            "obra":             obra_tr,
                            "observacao":       obs_tr,
                            "status":           "ATIVO",
                            "criado_por":       st.session_state.usuario_logado,
                        }
                        ok = insert_data("transferencias_tanque", dados_tr)
                        if ok:
                            st.success("✅ Transferência registrada com sucesso!")
                            time.sleep(1); st.rerun()

    with tab_hist:
        if df_tr.empty:
            st.info("Nenhuma transferência registrada.")
        else:
            df_tr = df_tr.sort_values("data", ascending=False).fillna("")
            fh1, fh2 = st.columns(2)
            fh_di = fh1.date_input("De",  value=date.today().replace(day=1), key="fh_di_tr")
            fh_df = fh2.date_input("Até", value=date.today(),                key="fh_df_tr")
            df_tr["data_dt"] = pd.to_datetime(df_tr["data"], errors="coerce").dt.date
            df_tr_fil = df_tr[(df_tr["data_dt"] >= fh_di) & (df_tr["data_dt"] <= fh_df)]

            cols_show = [c for c in ["data", "tanque_origem", "caminhao_tanque", "motorista",
                                     "produto", "quantidade", "valor_unitario", "total", "obra",
                                     "observacao", "status"] if c in df_tr_fil.columns]
            st.dataframe(df_tr_fil[cols_show], use_container_width=True)

            # Exportar
            if not df_tr_fil.empty:
                xl_tr = gerar_excel_limpo(df_tr_fil[cols_show], "Transferências")
                st.download_button(
                    "📥 Exportar Excel",
                    data=xl_tr,
                    file_name=f"Transferencias_{fh_di}_{fh_df}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )


# ════════════════════════════════════════════════════════════════════
# 3 · TANQUES / ESTOQUE
# ════════════════════════════════════════════════════════════════════
elif menu == "🛢️ Tanques / Estoque":
    st.markdown("## 🛢️ Gestão de Tanques e Comboios")

    df_t    = get_data("tanques")
    df_f    = get_data("fornecedores")
    df_ent  = get_data("entradas_tanque")
    df_sai  = get_data("abastecimentos")
    df_transf = get_data("transferencias_tanque")

    # Filtra apenas ativos
    if not df_sai.empty and "status" in df_sai.columns:
        df_sai = df_sai[df_sai["status"] == "ATIVO"]
    if not df_transf.empty and "status" in df_transf.columns:
        df_transf = df_transf[df_transf["status"] == "ATIVO"]

    # ── PAINEL DE SALDO ───────────────────────────────────────────────
    if not df_t.empty:
        st.markdown("### 📊 Visão Geral do Estoque")
        tanque_selecionado = st.selectbox("Selecione o Tanque/Comboio:", df_t["nome"].tolist())

        df_ent_filtrado   = df_ent[df_ent["nome_tanque"] == tanque_selecionado]   if not df_ent.empty   else pd.DataFrame()
        df_sai_filtrado   = df_sai[(df_sai.get("origem", "") == "Tanque Interno") & (df_sai["nome_tanque"] == tanque_selecionado)] if not df_sai.empty and "nome_tanque" in df_sai.columns else pd.DataFrame()
        df_transf_filtrado = df_transf[df_transf["tanque_origem"] == tanque_selecionado] if not df_transf.empty and "tanque_origem" in df_transf.columns else pd.DataFrame()

        tot_ent   = pd.to_numeric(df_ent_filtrado["quantidade"],   errors="coerce").fillna(0).sum()   if not df_ent_filtrado.empty   and "quantidade" in df_ent_filtrado.columns   else 0.0
        tot_sai   = pd.to_numeric(df_sai_filtrado["quantidade"],   errors="coerce").fillna(0).sum()   if not df_sai_filtrado.empty   and "quantidade" in df_sai_filtrado.columns   else 0.0
        tot_transf = pd.to_numeric(df_transf_filtrado["quantidade"], errors="coerce").fillna(0).sum() if not df_transf_filtrado.empty and "quantidade" in df_transf_filtrado.columns else 0.0

        saldo_atual = tot_ent - tot_sai - tot_transf

        c_ent, c_sai, c_ct, c_sal = st.columns(4)
        c_ent.metric("📥 Total de Entradas",          f"{tot_ent:,.1f} L")
        c_sai.metric("📤 Saídas Diretas",             f"{tot_sai:,.1f} L")
        c_ct.metric("🚛 Saídas p/ Caminhão-Tanque",   f"{tot_transf:,.1f} L")
        c_sal.metric("⛽ SALDO ATUAL",                 f"{saldo_atual:,.1f} L")
        st.markdown("---")

    # ── ABAS ─────────────────────────────────────────────────────────
    tab1, tab2 = st.tabs(["➕ Lançar Entrada de Combustível", "📋 Cadastrar Novo Tanque"])

    with tab1:
        if df_t.empty:
            st.warning("⚠️ Cadastre um tanque primeiro.")
        else:
            with st.form("f_ent_t", clear_on_submit=True):
                c1, c2 = st.columns(2)
                d_e  = c1.date_input("Data", value=date.today())
                nf_e = c2.text_input("NF / Documento")
                c3, c4 = st.columns(2)
                forn = c3.selectbox("Fornecedor (Distribuidora)", df_f["nome"].tolist() if not df_f.empty else ["Sem cadastro"])
                tanq = c4.selectbox("Tanque Destino", df_t["nome"].tolist())
                c5, c6, c7 = st.columns(3)
                prod = c5.selectbox("Produto", ["Diesel S10", "Diesel S500", "Gasolina Comum"])
                qtd  = c6.number_input("Quantidade (Litros)", min_value=0.0)
                vunt = c7.number_input("Valor Unitário (R$)", min_value=0.0)
                obs_e = st.text_input("Observação")
                if st.form_submit_button("📥 Registrar Entrada", use_container_width=True):
                    if qtd <= 0:
                        st.error("⚠️ Quantidade inválida.")
                    else:
                        ok = insert_data("entradas_tanque", {
                            "data":           str(d_e),
                            "numero_ficha":   nf_e,
                            "fornecedor":     forn,
                            "nome_tanque":    tanq,
                            "combustivel":    prod,
                            "quantidade":     qtd,
                            "valor_unitario": vunt,
                            "total":          qtd * vunt,
                            "observacao":     obs_e,
                            "criado_por":     st.session_state.usuario_logado,
                        })
                        if ok:
                            st.success("✅ Entrada salva!")
                            time.sleep(1); st.rerun()

            if not df_ent.empty:
                st.subheader("Últimas Entradas")
                df_e_rec = df_ent.sort_values("data", ascending=False).head(10).fillna("")
                cols_ent = ["data", "numero_ficha", "fornecedor", "nome_tanque", "combustivel",
                            "quantidade", "valor_unitario", "total", "criado_por"]
                cols_presentes = [c for c in cols_ent if c in df_e_rec.columns]
                st.dataframe(df_e_rec[cols_presentes], use_container_width=True)

    with tab2:
        with st.form("f_t", clear_on_submit=True):
            c1, c2 = st.columns(2)
            nm_t = c1.text_input("Nome do Tanque/Comboio")
            cap  = c2.number_input("Capacidade Máxima (Litros)", min_value=0.0)
            if st.form_submit_button("💾 Salvar Tanque", use_container_width=True):
                if nm_t:
                    ok = insert_data("tanques", {
                        "nome":       nm_t.upper(),
                        "capacidade": cap,
                        "criado_por": st.session_state.usuario_logado,
                    })
                    if ok:
                        st.success("✅ Tanque salvo!")
                        time.sleep(1); st.rerun()
                else:
                    st.error("⚠️ Preencha o nome.")
        if not df_t.empty:
            st.subheader("Tanques Cadastrados")
            for _, r in df_t.iterrows():
                cc1, cc2 = st.columns([4, 1])
                saldo_t = calcular_saldo(r["nome"])
                cc1.write(f"**{r['nome']}** — Capacidade: {r.get('capacidade', 0):.0f} L | Saldo: {saldo_t:,.1f} L")
                if cc2.button("❌", key=f"d_t_{r['id']}"):
                    if delete_data("tanques", r["id"]):
                        st.rerun()


# ════════════════════════════════════════════════════════════════════
# 4 · BOLETIM DE TRANSPORTE E PRODUÇÃO
# ════════════════════════════════════════════════════════════════════
elif menu == "🚚 Boletim de Transporte":
    st.markdown("## 🚚 Boletim Diário de Produção")
    df_v = get_data("veiculos")
    if df_v.empty:
        st.warning("⚠️ Cadastre veículos primeiro.")
        st.stop()

    obras_lista = lista_obras()

    with st.form("f_prod", clear_on_submit=True):
        st.markdown("#### 👤 Dados Operacionais")
        c1, c2, c3 = st.columns(3)
        dt_p = c1.date_input("Data do Boletim", value=date.today())
        pref = c2.selectbox("Veículo / Equipamento", df_v["prefixo"].tolist())
        v_info = df_v[df_v["prefixo"] == pref].iloc[0]
        mot  = c3.text_input("Motorista / Operador", value=v_info.get("motorista", ""))

        st.markdown("#### 🏗️ Obra e Rota")
        c4, c5, c6 = st.columns(3)
        if obras_lista:
            obra_bol = c4.selectbox("Obra / Projeto", obras_lista)
        else:
            obra_bol = c4.text_input("Obra / Projeto")
        origem_rota  = c5.text_input("Origem / Jazida (Ex: Usina, Pedreira, Base)")
        destino_rota = c6.text_input("Destino / Trecho de Aplicação")

        st.markdown("#### 🛣️ Tipo de Operação e Produção")
        c7, c8, c9, c10 = st.columns(4)
        op_tipo  = c7.selectbox("Tipo de Operação", ["Transporte de Massa/CBUQ", "Transporte de Fresado", "Terraplanagem", "Venda de Massa", "Ocioso/Manutenção"])
        km_s     = c8.number_input("KM Inicial",  min_value=0.0)
        km_c     = c9.number_input("KM Final",    min_value=0.0)
        carradas = c10.number_input("Nº de Carradas/Viagens", min_value=0, step=1)

        c11, c12 = st.columns(2)
        ton = c11.number_input("Total Toneladas", min_value=0.0)
        obs_p = c12.text_input("Observações Gerais")

        st.markdown("#### ⛽ Abastecimento na Viagem (Opcional)")
        c13, c14, c15 = st.columns(3)
        houve_abast      = c13.checkbox("Houve abastecimento externo nesta rota?")
        litros_rota      = c14.number_input("Litros abastecidos",  min_value=0.0) if houve_abast else 0.0
        valor_abast_rota = c15.number_input("Valor Total (R$)",    min_value=0.0) if houve_abast else 0.0

        if st.form_submit_button("💾 Salvar Boletim Diário", use_container_width=True):
            if op_tipo != "Ocioso/Manutenção" and carradas <= 0:
                st.error("⚠️ Insira a quantidade de viagens.")
            else:
                ok = insert_data("producao", {
                    "data":                str(dt_p),
                    "prefixo":             pref,
                    "motorista":           mot.upper(),
                    "obra":                obra_bol,
                    "tipo_operacao":       op_tipo,
                    "origem":              origem_rota.upper(),
                    "destino":             destino_rota.upper(),
                    "local_aplicacao":     destino_rota.upper(),
                    "km_saida":            km_s,
                    "km_chegada":          km_c,
                    "carradas":            carradas,
                    "toneladas":           ton,
                    "abastecimento_litros": litros_rota,
                    "abastecimento_valor": valor_abast_rota,
                    "observacao":          obs_p,
                    "criado_por":          st.session_state.usuario_logado,
                })
                if ok:
                    st.success("✅ Boletim salvo!")
                    time.sleep(1); st.rerun()

    df_bol = get_data("producao")
    if not df_bol.empty:
        st.divider()
        st.subheader("📋 Últimos Boletins Registrados")
        df_br = df_bol.sort_values("data", ascending=False).head(20).fillna("")
        colunas_bol = ["data", "prefixo", "motorista", "obra", "tipo_operacao",
                       "origem", "destino", "carradas", "toneladas",
                       "abastecimento_litros", "km_saida", "km_chegada"]
        colunas_presentes = [c for c in colunas_bol if c in df_br.columns]
        st.dataframe(df_br[colunas_presentes], use_container_width=True)


# ════════════════════════════════════════════════════════════════════
# 5 · FROTA E EQUIPAMENTOS
# ════════════════════════════════════════════════════════════════════
elif menu == "🚜 Frota e Equipamentos":
    st.markdown("## 🚜 Gestão de Frota")
    df_v = get_data("veiculos")

    # Proteção de colunas
    if not df_v.empty:
        for col in ["categoria", "tipo_veiculo", "tipo_combustivel_padrao", "motorista", "placa"]:
            if col not in df_v.columns:
                df_v[col] = ""

    with st.expander("➕ CADASTRAR NOVO VEÍCULO / EQUIPAMENTO", expanded=True):
        with st.form("f_v", clear_on_submit=True):
            c1, c2, c3 = st.columns(3)
            pref      = c1.text_input("Código / Prefixo (Ex: CA-01)")
            plc       = c2.text_input("Placa")
            categoria = c3.selectbox("Categoria", ["Veículo", "Equipamento"])

            c4, c5, c6 = st.columns(3)
            mot       = c4.text_input("Motorista / Operador Fixo")
            comb      = c5.selectbox("Combustível Padrão", ["Diesel S10", "Diesel S500", "Gasolina Comum"])
            tipo_veic = c6.selectbox(
                "Tipo de Veículo",
                ["Veículo", "Equipamento", "Caminhão-Tanque"],
                help="Selecione 'Caminhão-Tanque' para veículos que abastecem outros em campo."
            )

            if st.form_submit_button("💾 Salvar", use_container_width=True):
                if pref:
                    ok = insert_data("veiculos", {
                        "prefixo":                pref.upper(),
                        "placa":                  plc.upper(),
                        "categoria":              categoria,
                        "motorista":              mot.upper(),
                        "tipo_combustivel_padrao": comb,
                        "tipo_veiculo":           tipo_veic,
                    })
                    if ok:
                        st.success("✅ Salvo com sucesso!")
                        st.rerun()
                else:
                    st.error("⚠️ Prefixo é obrigatório.")

    if not df_v.empty:
        st.divider()

        # Separar por tipo
        tab_veic, tab_ct = st.tabs(["🚜 Veículos e Equipamentos", "🚛 Caminhões-Tanque"])

        with tab_veic:
            df_normais = df_v[df_v.get("tipo_veiculo", pd.Series(dtype=str)) != "Caminhão-Tanque"] if "tipo_veiculo" in df_v.columns else df_v
            if df_normais.empty:
                st.info("Nenhum veículo/equipamento cadastrado.")
            else:
                for _, r in df_normais.iterrows():
                    cc1, cc2 = st.columns([5, 1])
                    cc1.markdown(
                        f"**{r.get('prefixo', '')}** | "
                        f"{r.get('tipo_veiculo', r.get('categoria', '-'))} | "
                        f"Placa: {r.get('placa', '')} | "
                        f"Operador: {r.get('motorista', '')} | "
                        f"Combustível: {r.get('tipo_combustivel_padrao', '')}"
                    )
                    if cc2.button("❌ Excluir", key=f"d_v_{r.get('id', 'x')}"):
                        if delete_data("veiculos", r.get("id")):
                            st.rerun()

        with tab_ct:
            df_ct = df_v[df_v.get("tipo_veiculo", pd.Series(dtype=str)) == "Caminhão-Tanque"] if "tipo_veiculo" in df_v.columns else pd.DataFrame()
            if df_ct.empty:
                st.info("Nenhum caminhão-tanque cadastrado. Cadastre acima e selecione 'Caminhão-Tanque'.")
            else:
                for _, r in df_ct.iterrows():
                    cc1, cc2 = st.columns([5, 1])
                    cc1.markdown(
                        f"🚛 **{r.get('prefixo', '')}** | "
                        f"Placa: {r.get('placa', '')} | "
                        f"Motorista: {r.get('motorista', '')} | "
                        f"Combustível: {r.get('tipo_combustivel_padrao', '')}"
                    )
                    if cc2.button("❌ Excluir", key=f"d_ct_{r.get('id', 'x')}"):
                        if delete_data("veiculos", r.get("id")):
                            st.rerun()


# ════════════════════════════════════════════════════════════════════
# 6 · OBRAS
# ════════════════════════════════════════════════════════════════════
elif menu == "🏗️ Obras":
    st.markdown("## 🏗️ Gestão de Obras e Projetos")
    st.markdown("""
    <div class='banner-info'>
    ℹ️ Cadastre aqui todas as obras/projetos. Elas estarão disponíveis como opção em
    <strong>Abastecimentos</strong>, <strong>Transferências</strong> e <strong>Boletins de Transporte</strong>,
    garantindo rastreabilidade completa do combustível por obra.
    </div>
    """, unsafe_allow_html=True)

    df_o = get_data("obras")

    with st.expander("➕ CADASTRAR NOVA OBRA", expanded=True):
        with st.form("f_obra", clear_on_submit=True):
            co1, co2, co3 = st.columns(3)
            nome_obra   = co1.text_input("Nome da Obra / Projeto")
            codigo_obra = co2.text_input("Código / ART")
            status_obra = co3.selectbox("Status", ["Ativa", "Pausada", "Encerrada"])
            co4, co5 = st.columns(2)
            local_obra = co4.text_input("Município / Localização")
            resp_obra  = co5.text_input("Responsável Técnico")
            obs_obra   = st.text_input("Observações")

            if st.form_submit_button("💾 Salvar Obra", use_container_width=True):
                if nome_obra:
                    ok = insert_data("obras", {
                        "nome":        nome_obra.upper(),
                        "codigo":      codigo_obra.upper(),
                        "status":      status_obra,
                        "local":       local_obra.upper(),
                        "responsavel": resp_obra.upper(),
                        "observacao":  obs_obra,
                        "criado_por":  st.session_state.usuario_logado,
                    })
                    if ok:
                        st.success("✅ Obra salva!")
                        time.sleep(1); st.rerun()
                else:
                    st.error("⚠️ Nome da obra é obrigatório.")

    if not df_o.empty:
        st.divider()
        st.subheader("📋 Obras Cadastradas")

        # KPIs por obra (consumo de combustível)
        df_ab_o = get_data("abastecimentos")
        if not df_ab_o.empty and "status" in df_ab_o.columns:
            df_ab_o = df_ab_o[df_ab_o["status"] == "ATIVO"]

        for _, r in df_o.iterrows():
            nome_o = r.get("nome", "")
            status_o = r.get("status", "Ativa")
            cor_status = "banner-ok" if status_o == "Ativa" else ("banner-low" if status_o == "Pausada" else "banner-err")

            with st.expander(f"🏗️ {nome_o} | {r.get('codigo', '')} | {status_o}"):
                oc1, oc2, oc3 = st.columns(3)
                oc1.write(f"**Local:** {r.get('local', '-')}")
                oc2.write(f"**Responsável:** {r.get('responsavel', '-')}")
                oc3.write(f"**Obs:** {r.get('observacao', '-')}")

                # Consumo total da obra
                if not df_ab_o.empty and "obra" in df_ab_o.columns:
                    df_obra_ab = df_ab_o[df_ab_o["obra"] == nome_o]
                    total_l = pd.to_numeric(df_obra_ab.get("quantidade", 0), errors="coerce").sum()
                    total_r = pd.to_numeric(df_obra_ab.get("total",      0), errors="coerce").sum()
                    oc4, oc5, oc6 = st.columns(3)
                    oc4.metric("⛽ Litros consumidos", f"{total_l:,.1f} L")
                    oc5.metric("💰 Gasto total",       f"R$ {total_r:,.2f}")
                    oc6.metric("📋 Abastecimentos",    len(df_obra_ab))

                col_edit, col_del = st.columns(2)
                # Alterar status
                novo_status = col_edit.selectbox(
                    "Alterar Status",
                    ["Ativa", "Pausada", "Encerrada"],
                    index=["Ativa", "Pausada", "Encerrada"].index(status_o) if status_o in ["Ativa", "Pausada", "Encerrada"] else 0,
                    key=f"st_obra_{r['id']}"
                )
                if col_edit.button("💾 Atualizar Status", key=f"upd_obra_{r['id']}"):
                    if update_data("obras", r["id"], {"status": novo_status}):
                        st.success("Status atualizado!")
                        time.sleep(1); st.rerun()
                if col_del.button("❌ Excluir Obra", key=f"del_obra_{r['id']}"):
                    if delete_data("obras", r["id"]):
                        st.rerun()


# ════════════════════════════════════════════════════════════════════
# 7 · FORNECEDORES
# ════════════════════════════════════════════════════════════════════
elif menu == "🏪 Fornecedores":
    st.markdown("## 🏪 Postos e Distribuidoras")
    df_f = get_data("fornecedores")

    with st.expander("➕ CADASTRAR NOVO FORNECEDOR", expanded=True):
        with st.form("f_f", clear_on_submit=True):
            c1, c2 = st.columns(2)
            nm = c1.text_input("Nome Fantasia (Aparece no App)")
            rz = c2.text_input("Razão Social")
            c2b, c2c = st.columns(2)
            cnpj = c2b.text_input("CNPJ")
            tel  = c2c.text_input("Telefone / Contato")
            st.markdown("**DADOS BANCÁRIOS / EXPORTAÇÃO**")
            c3, c4, c5 = st.columns(3)
            banco = c3.text_input("Banco")
            ag    = c4.text_input("Agência")
            cta   = c5.text_input("Conta")
            c6, c7 = st.columns(2)
            tipo_c = c6.selectbox("Tipo Conta", ["Corrente", "Poupança"])
            pix    = c7.text_input("Chave PIX")
            st.markdown("**PREÇOS CONTRATUAIS (Opcional)**")
            c8, c9 = st.columns(2)
            p_d = c8.number_input("Preço Diesel Acordado (R$)",    min_value=0.0)
            p_g = c9.number_input("Preço Gasolina Acordado (R$)",  min_value=0.0)
            if st.form_submit_button("💾 Salvar Fornecedor", use_container_width=True):
                if nm:
                    ok = insert_data("fornecedores", {
                        "nome":         nm.upper(),
                        "razao_social": rz.upper(),
                        "cnpj":         cnpj,
                        "telefone":     tel,
                        "banco":        banco,
                        "agencia":      ag,
                        "conta":        cta,
                        "tipo_conta":   tipo_c,
                        "pix":          pix,
                        "preco_diesel": p_d,
                        "preco_gasolina": p_g,
                        "criado_por":   st.session_state.usuario_logado,
                    })
                    if ok:
                        st.success("✅ Salvo!")
                        time.sleep(1); st.rerun()
                else:
                    st.error("⚠️ Nome Fantasia é obrigatório.")

    if not df_f.empty:
        st.divider()
        st.subheader("📋 Fornecedores Cadastrados")
        for _, r in df_f.iterrows():
            cc1, cc2 = st.columns([5, 1])
            cc1.markdown(
                f"**{r['nome']}** | "
                f"CNPJ: {r.get('cnpj', '-')} | "
                f"Banco: {r.get('banco', '')} Ag: {r.get('agencia', '')} Cc: {r.get('conta', '')} | "
                f"PIX: {r.get('pix', '')}"
            )
            if cc2.button("❌ Excluir", key=f"d_f_{r['id']}"):
                if delete_data("fornecedores", r["id"]):
                    st.rerun()


# ════════════════════════════════════════════════════════════════════
# 8 · RELATÓRIOS E FECHAMENTOS
# ════════════════════════════════════════════════════════════════════
elif menu == "📋 Relatórios e Fechamentos":
    st.markdown("## 📋 Central de Relatórios")

    aba1, aba2, aba3, aba4 = st.tabs([
        "⛽ Relatório de Saídas",
        "🛢️ Fechamento de Tanques",
        "📉 Resumo Gerencial / Produção",
        "🔗 Rastreabilidade por Obra",
    ])

    # ── ABA 1: SAÍDAS ────────────────────────────────────────────────
    with aba1:
        st.markdown("#### Gerar Fechamento de Posto Externo")
        df_a = get_data("abastecimentos")
        df_f = get_data("fornecedores")

        if not df_a.empty and "status" in df_a.columns:
            df_a = df_a[df_a["status"] == "ATIVO"]

        c1, c2, c3, c4 = st.columns(4)
        dt_i     = c1.date_input("Início (Saídas)", value=date.today().replace(day=1))
        dt_f_s   = c2.date_input("Fim (Saídas)",    value=date.today())
        forn_sel = c3.selectbox("Filtrar Fornecedor", ["TODOS"] + (df_f["nome"].tolist() if not df_f.empty else []))
        obras_rel = lista_obras(incluir_todas=True)
        obra_rel  = c4.selectbox("Filtrar Obra", obras_rel if obras_rel else ["TODAS"])

        if st.button("🔍 Filtrar Saídas"):
            if df_a.empty:
                st.warning("Sem dados.")
            else:
                df_a["data_dt"] = pd.to_datetime(df_a["data"], errors="coerce").dt.date
                df_filtro = df_a[(df_a["data_dt"] >= dt_i) & (df_a["data_dt"] <= dt_f_s)]
                if forn_sel != "TODOS":
                    df_filtro = df_filtro[df_filtro["fornecedor"] == forn_sel]
                if obra_rel not in ("TODAS", "") and "obra" in df_filtro.columns:
                    df_filtro = df_filtro[df_filtro["obra"] == obra_rel]
                df_filtro = df_filtro.sort_values("data")

                st.write(f"Encontrados **{len(df_filtro)}** registros.")
                if not df_filtro.empty:
                    # Totais rápidos
                    tot_l = pd.to_numeric(df_filtro.get("quantidade", 0), errors="coerce").sum()
                    tot_r = pd.to_numeric(df_filtro.get("total",      0), errors="coerce").sum()
                    m1, m2 = st.columns(2)
                    m1.metric("⛽ Total Litros", f"{tot_l:,.1f} L")
                    m2.metric("💰 Total Gasto",  f"R$ {tot_r:,.2f}")

                    per_str    = f"{dt_i.strftime('%d/%m/%Y')} a {dt_f_s.strftime('%d/%m/%Y')}"
                    obra_padrao = obra_rel if obra_rel not in ("TODAS", "") else "COPA ENGENHARIA"
                    dados_forn  = {}
                    if forn_sel != "TODOS" and not df_f.empty:
                        fs = df_f[df_f["nome"] == forn_sel]
                        if not fs.empty:
                            dados_forn = fs.iloc[0].to_dict()

                    c4b, c5b, c6b = st.columns(3)
                    xl_copa = gerar_excel_copa(df_filtro, dados_forn, per_str, obra_padrao, forn_sel if forn_sel != "TODOS" else "GERAL")
                    c4b.download_button("📥 Excel Padrão COPA", data=xl_copa,
                                        file_name=f"Abastecimentos_{forn_sel}_{dt_i}.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        use_container_width=True)

                    dados_pdf = {"FORNECEDOR": forn_sel if forn_sel != "TODOS" else "GERAL", "PERÍODO": per_str}
                    if dados_forn:
                        dados_pdf["CNPJ/CPF"] = dados_forn.get("cnpj", "")
                        dados_pdf["BANCO"]    = dados_forn.get("banco", "")
                        dados_pdf["AGÊNCIA"]  = dados_forn.get("agencia", "")
                        dados_pdf["CONTA"]    = dados_forn.get("conta", "")
                        dados_pdf["PIX"]      = dados_forn.get("pix", "")
                    pdf_bytes = gerar_pdf(
                        df_filtro, "SAIDAS",
                        "COPA ENGENHARIA LTDA", "DEPARTAMENTO DE EQUIPAMENTOS",
                        f"PERÍODO: {per_str}", dados_pdf,
                        f"RELATÓRIO DE ABASTECIMENTOS - {forn_sel if forn_sel != 'TODOS' else 'GERAL'}"
                    )
                    c5b.download_button("📄 PDF Timbrado", data=pdf_bytes,
                                        file_name=f"Relatorio_{forn_sel}_{dt_i}.pdf",
                                        mime="application/pdf", use_container_width=True)

                    xl_limpo = gerar_excel_limpo(df_filtro, "Saídas")
                    c6b.download_button("📊 Excel Tabela Limpa", data=xl_limpo,
                                        file_name="Tabela_Saidas.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        use_container_width=True)

                    st.dataframe(df_filtro.drop(columns=["data_dt"], errors="ignore"), use_container_width=True)

    # ── ABA 2: FECHAMENTO DE TANQUES ─────────────────────────────────
    with aba2:
        st.markdown("#### Fechamento Físico de Tanques (Entradas vs Saídas vs Transferências)")
        df_ent  = get_data("entradas_tanque")
        df_t    = get_data("tanques")
        df_ab2  = get_data("abastecimentos")
        df_tr2  = get_data("transferencias_tanque")

        if not df_ab2.empty and "status" in df_ab2.columns:
            df_ab2 = df_ab2[df_ab2["status"] == "ATIVO"]
        if not df_tr2.empty and "status" in df_tr2.columns:
            df_tr2 = df_tr2[df_tr2["status"] == "ATIVO"]

        c1, c2, c3 = st.columns(3)
        di_t    = c1.date_input("Data Início (Tanque)", value=date.today().replace(day=1))
        df_tq   = c2.date_input("Data Fim (Tanque)",    value=date.today())
        tanq_sel = c3.selectbox("Selecionar Tanque", df_t["nome"].tolist() if not df_t.empty else ["Sem cadastro"])

        if st.button("🔍 Gerar Fechamento de Tanque"):
            per_t = f"{di_t.strftime('%d/%m/%Y')} a {df_tq.strftime('%d/%m/%Y')}"
            ent_f = pd.DataFrame(); sai_f = pd.DataFrame(); transf_f = pd.DataFrame()

            if not df_ent.empty:
                df_ent["data_dt"] = pd.to_datetime(df_ent["data"], errors="coerce").dt.date
                ent_f = df_ent[
                    (df_ent["data_dt"] >= di_t) &
                    (df_ent["data_dt"] <= df_tq) &
                    (df_ent["nome_tanque"] == tanq_sel)
                ]
            if not df_ab2.empty and "nome_tanque" in df_ab2.columns:
                sai_f = df_ab2[(df_ab2["origem"] == "Tanque Interno") & (df_ab2["nome_tanque"] == tanq_sel)]
                if not sai_f.empty:
                    sai_f["data_dt"] = pd.to_datetime(sai_f["data"], errors="coerce").dt.date
                    sai_f = sai_f[(sai_f["data_dt"] >= di_t) & (sai_f["data_dt"] <= df_tq)]
            if not df_tr2.empty and "tanque_origem" in df_tr2.columns:
                transf_f = df_tr2[df_tr2["tanque_origem"] == tanq_sel]
                if not transf_f.empty:
                    transf_f["data_dt"] = pd.to_datetime(transf_f["data"], errors="coerce").dt.date
                    transf_f = transf_f[(transf_f["data_dt"] >= di_t) & (transf_f["data_dt"] <= df_tq)]

            if ent_f.empty and sai_f.empty and transf_f.empty:
                st.warning("Nenhum movimento no período.")
            else:
                tot_e  = pd.to_numeric(ent_f.get("quantidade",   0), errors="coerce").sum()   if not ent_f.empty   else 0
                tot_s  = pd.to_numeric(sai_f.get("quantidade",   0), errors="coerce").sum()   if not sai_f.empty   else 0
                tot_tr = pd.to_numeric(transf_f.get("quantidade", 0), errors="coerce").sum()  if not transf_f.empty else 0
                m1, m2, m3, m4 = st.columns(4)
                m1.metric("📥 Entradas",          f"{tot_e:,.1f} L")
                m2.metric("📤 Saídas Diretas",    f"{tot_s:,.1f} L")
                m3.metric("🚛 Transf. Caminhão",  f"{tot_tr:,.1f} L")
                m4.metric("⛽ Saldo Período",      f"{tot_e - tot_s - tot_tr:,.1f} L")

                xl_t = gerar_excel_tanque(ent_f, sai_f, transf_f, tanq_sel, per_t, "COPA ENGENHARIA")
                c4t, c5t = st.columns(2)
                c4t.download_button("📥 Baixar Excel do Tanque", data=xl_t,
                                    file_name=f"Tanque_{tanq_sel}_{di_t}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True)

                # Montar movimentos para PDF
                movs = []
                if not ent_f.empty:
                    for _, r in ent_f.iterrows():
                        movs.append({"data": r.get("data", ""), "tipo": "ENTRADA",
                                     "numero_ficha": r.get("numero_ficha", ""),
                                     "motorista_forn": r.get("fornecedor", ""),
                                     "produto": r.get("combustivel", ""),
                                     "qtd_entrada": r.get("quantidade", 0), "qtd_saida": 0,
                                     "valor_unitario": r.get("valor_unitario", 0),
                                     "total": r.get("total", 0),
                                     "obra": r.get("obra", ""),
                                     "observacao": r.get("observacao", "")})
                if not sai_f.empty:
                    for _, r in sai_f.iterrows():
                        movs.append({"data": r.get("data", ""), "tipo": "SAÍDA DIRETA",
                                     "numero_ficha": r.get("numero_ficha", ""),
                                     "placa": r.get("placa", ""), "prefixo": r.get("prefixo", ""),
                                     "motorista_forn": r.get("motorista", ""),
                                     "produto": r.get("tipo_combustivel", ""),
                                     "horimetro": r.get("horimetro", ""),
                                     "qtd_entrada": 0, "qtd_saida": r.get("quantidade", 0),
                                     "valor_unitario": r.get("valor_unitario", 0),
                                     "total": r.get("total", 0),
                                     "obra": r.get("obra", ""),
                                     "observacao": r.get("observacao", "")})
                if not transf_f.empty:
                    for _, r in transf_f.iterrows():
                        movs.append({"data": r.get("data", ""), "tipo": "TRANSF. CAMINHÃO",
                                     "numero_ficha": r.get("numero_ficha", ""),
                                     "placa": r.get("placa", ""),
                                     "prefixo": r.get("caminhao_tanque", ""),
                                     "motorista_forn": r.get("motorista", ""),
                                     "produto": r.get("produto", ""),
                                     "qtd_entrada": 0, "qtd_saida": r.get("quantidade", 0),
                                     "valor_unitario": r.get("valor_unitario", 0),
                                     "total": r.get("total", 0),
                                     "obra": r.get("obra", ""),
                                     "observacao": r.get("observacao", "")})

                df_movs = pd.DataFrame(movs).sort_values("data")
                pdf_t = gerar_pdf(df_movs, "TANQUE",
                                  "COPA ENGENHARIA LTDA", "CONTROLE DE ESTOQUE",
                                  f"PERÍODO: {per_t}", {"TANQUE": tanq_sel},
                                  f"FECHAMENTO FÍSICO DE TANQUE - {tanq_sel}")
                c5t.download_button("📄 Baixar PDF do Tanque", data=pdf_t,
                                    file_name=f"Tanque_{tanq_sel}_{di_t}.pdf",
                                    mime="application/pdf", use_container_width=True)

    # ── ABA 3: RESUMO GERENCIAL ───────────────────────────────────────
    with aba3:
        st.markdown("#### Exportar Boletins de Produção")
        df_prod = get_data("producao")
        di_p, df_p = st.columns(2)
        d1 = di_p.date_input("De (Produção)",  value=date.today().replace(day=1))
        d2 = df_p.date_input("Até (Produção)", value=date.today())
        if st.button("📊 Extrair Tabela de Produção"):
            if not df_prod.empty:
                df_prod["data_dt"] = pd.to_datetime(df_prod["data"], errors="coerce").dt.date
                df_pf = df_prod[(df_prod["data_dt"] >= d1) & (df_prod["data_dt"] <= d2)]
                if not df_pf.empty:
                    xl_p = gerar_excel_limpo(df_pf.drop(columns=["data_dt"]), "Producao")
                    st.download_button("📥 Baixar Excel Produção", data=xl_p,
                                       file_name=f"Producao_{d1}_a_{d2}.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    st.dataframe(df_pf.drop(columns=["data_dt"]), use_container_width=True)
                else:
                    st.info("Nenhum lançamento no período.")
            else:
                st.info("Nenhum lançamento no banco de dados.")

    # ── ABA 4: RASTREABILIDADE POR OBRA ──────────────────────────────
    with aba4:
        st.markdown("#### 🔗 Rastreabilidade Completa por Obra")
        st.markdown("""
        <div class='banner-info'>
        ℹ️ Visualize o fluxo completo de combustível por obra:
        <strong>Compra → Tanque → Caminhão-Tanque → Veículo → Obra</strong>
        </div>
        """, unsafe_allow_html=True)

        obras_r = lista_obras()
        if not obras_r:
            st.warning("⚠️ Nenhuma obra cadastrada. Cadastre obras na aba 🏗️ Obras.")
        else:
            cr1, cr2, cr3 = st.columns(3)
            obra_rastr = cr1.selectbox("Selecionar Obra", obras_r)
            dr_i = cr2.date_input("De",  value=date.today().replace(day=1), key="dr_i")
            dr_f = cr3.date_input("Até", value=date.today(),                key="dr_f")

            if st.button("🔍 Gerar Rastreabilidade"):
                # Abastecimentos diretos da obra
                df_ab_r = get_data("abastecimentos")
                if not df_ab_r.empty and "status" in df_ab_r.columns:
                    df_ab_r = df_ab_r[df_ab_r["status"] == "ATIVO"]
                if not df_ab_r.empty and "obra" in df_ab_r.columns:
                    df_ab_r["data_dt"] = pd.to_datetime(df_ab_r["data"], errors="coerce").dt.date
                    df_ab_r = df_ab_r[
                        (df_ab_r["obra"] == obra_rastr) &
                        (df_ab_r["data_dt"] >= dr_i) &
                        (df_ab_r["data_dt"] <= dr_f)
                    ]

                # Transferências de caminhão-tanque para a obra
                df_tr_r = get_data("transferencias_tanque")
                if not df_tr_r.empty and "status" in df_tr_r.columns:
                    df_tr_r = df_tr_r[df_tr_r["status"] == "ATIVO"]
                if not df_tr_r.empty and "obra" in df_tr_r.columns:
                    df_tr_r["data_dt"] = pd.to_datetime(df_tr_r["data"], errors="coerce").dt.date
                    df_tr_r = df_tr_r[
                        (df_tr_r["obra"] == obra_rastr) &
                        (df_tr_r["data_dt"] >= dr_i) &
                        (df_tr_r["data_dt"] <= dr_f)
                    ]

                # Boletins de produção da obra
                df_prod_r = get_data("producao")
                if not df_prod_r.empty and "obra" in df_prod_r.columns:
                    df_prod_r["data_dt"] = pd.to_datetime(df_prod_r["data"], errors="coerce").dt.date
                    df_prod_r = df_prod_r[
                        (df_prod_r["obra"] == obra_rastr) &
                        (df_prod_r["data_dt"] >= dr_i) &
                        (df_prod_r["data_dt"] <= dr_f)
                    ]

                # KPIs consolidados
                tot_l_ab  = pd.to_numeric(df_ab_r.get("quantidade", 0),   errors="coerce").sum() if not df_ab_r.empty else 0
                tot_r_ab  = pd.to_numeric(df_ab_r.get("total",      0),   errors="coerce").sum() if not df_ab_r.empty else 0
                tot_l_tr  = pd.to_numeric(df_tr_r.get("quantidade", 0),   errors="coerce").sum() if not df_tr_r.empty else 0
                tot_r_tr  = pd.to_numeric(df_tr_r.get("total",      0),   errors="coerce").sum() if not df_tr_r.empty else 0
                tot_ton   = pd.to_numeric(df_prod_r.get("toneladas", 0),  errors="coerce").sum() if not df_prod_r.empty else 0
                tot_viag  = pd.to_numeric(df_prod_r.get("carradas",  0),  errors="coerce").sum() if not df_prod_r.empty else 0

                total_litros = tot_l_ab + tot_l_tr
                total_gasto  = tot_r_ab + tot_r_tr

                st.markdown(f"### 📊 Resumo — {obra_rastr}")
                km1, km2, km3, km4 = st.columns(4)
                km1.metric("⛽ Total Litros Consumidos", f"{total_litros:,.1f} L")
                km2.metric("💰 Gasto Total Combustível", f"R$ {total_gasto:,.2f}")
                km3.metric("🏗️ Toneladas Produzidas",   f"{tot_ton:,.1f} t")
                km4.metric("🚚 Viagens Realizadas",      int(tot_viag))

                custo_ton_r = total_gasto / tot_ton if tot_ton > 0 else 0
                km5, km6 = st.columns(2)
                km5.metric("💡 Custo Combustível / Ton", f"R$ {custo_ton_r:,.2f}")
                km6.metric("💡 Litros / Ton",            f"{(total_litros / tot_ton):,.2f} L" if tot_ton > 0 else "—")

                st.divider()

                # Detalhamento
                if not df_ab_r.empty:
                    st.markdown("##### ⛽ Abastecimentos Diretos (Posto Externo e Tanque)")
                    cols_ab = [c for c in ["data", "prefixo", "placa", "motorista", "fornecedor",
                                           "tipo_combustivel", "quantidade", "valor_unitario", "total",
                                           "origem", "horimetro", "observacao"] if c in df_ab_r.columns]
                    st.dataframe(df_ab_r[cols_ab], use_container_width=True)

                if not df_tr_r.empty:
                    st.markdown("##### 🚛 Transferências via Caminhão-Tanque")
                    cols_tr = [c for c in ["data", "tanque_origem", "caminhao_tanque", "motorista",
                                           "produto", "quantidade", "valor_unitario", "total",
                                           "observacao"] if c in df_tr_r.columns]
                    st.dataframe(df_tr_r[cols_tr], use_container_width=True)

                if not df_prod_r.empty:
                    st.markdown("##### 🚚 Boletins de Produção")
                    cols_prod = [c for c in ["data", "prefixo", "motorista", "tipo_operacao",
                                             "origem", "destino", "carradas", "toneladas",
                                             "km_saida", "km_chegada"] if c in df_prod_r.columns]
                    st.dataframe(df_prod_r[cols_prod], use_container_width=True)

                # Exportar rastreabilidade completa
                if not df_ab_r.empty or not df_tr_r.empty:
                    buf_rastr = io.BytesIO()
                    with pd.ExcelWriter(buf_rastr, engine="xlsxwriter") as wr:
                        if not df_ab_r.empty:
                            df_ab_r.drop(columns=["data_dt"], errors="ignore").to_excel(wr, index=False, sheet_name="Abastecimentos")
                        if not df_tr_r.empty:
                            df_tr_r.drop(columns=["data_dt"], errors="ignore").to_excel(wr, index=False, sheet_name="Transferencias")
                        if not df_prod_r.empty:
                            df_prod_r.drop(columns=["data_dt"], errors="ignore").to_excel(wr, index=False, sheet_name="Producao")
                    st.download_button(
                        "📥 Exportar Rastreabilidade Completa (Excel)",
                        data=buf_rastr.getvalue(),
                        file_name=f"Rastreabilidade_{obra_rastr}_{dr_i}_{dr_f}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )


# ════════════════════════════════════════════════════════════════════
# 9 · USUÁRIOS E ACESSOS
# ════════════════════════════════════════════════════════════════════
elif menu == "👥 Usuários e Acessos":
    if st.session_state.perfil_logado != "Admin":
        st.error("⛔ Acesso restrito.")
        st.stop()
    st.markdown("## 👥 Gestão de Usuários e Acessos")
    st.info("O campo **'Criado Por'** registra automaticamente quem fez cada lançamento.")

    with st.form("f_usr", clear_on_submit=True):
        c1, c2 = st.columns(2)
        nm_u = c1.text_input("Nome Completo")
        lg_u = c2.text_input("Login")
        c3, c4 = st.columns(2)
        sn_u = c3.text_input("Senha", type="password")
        pf_u = c4.selectbox("Perfil", ["Operador", "Admin"])
        if st.form_submit_button("💾 Criar Usuário", use_container_width=True):
            if nm_u and lg_u and sn_u:
                ok = insert_data("usuarios", {"nome": nm_u, "login": lg_u, "senha": sn_u, "perfil": pf_u})
                if ok:
                    st.success("✅ Usuário criado!")
                    st.rerun()
            else:
                st.error("⚠️ Preencha todos os campos.")

    df_u = get_data("usuarios")
    if not df_u.empty:
        st.divider()
        st.subheader("Usuários Cadastrados")
        for _, r in df_u.iterrows():
            cc1, cc2 = st.columns([5, 1])
            cc1.write(f"**{r.get('nome', '')}** ({r.get('login', '')}) — Nível: {r.get('perfil', '')}")
            if cc2.button("❌", key=f"del_u_{r['id']}"):
                if delete_data("usuarios", r["id"]):
                    st.rerun()
